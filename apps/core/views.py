"""API and page views for the automation testing module."""

from __future__ import annotations

from typing import Any

import base64
import binascii
import io
import json
import logging
import os
import signal
import subprocess
import sys
import threading
import time
from pathlib import Path
from urllib.parse import urlparse

import requests

from django.conf import settings
# Optional dependency: PyCryptodome for AES-CBC handling
try:  # pragma: no cover - environment dependent
    from Crypto.Cipher import AES as _AES
    from Crypto.Util.Padding import pad as _pad
    from Crypto.Util.Padding import unpad as _unpad
except ImportError:  # pragma: no cover - handled gracefully at runtime
    _AES = None
    _pad = None
    _unpad = None

from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Q, Count, Max
from django.http import HttpResponse
from django.http import Http404
from django.shortcuts import render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.csrf import ensure_csrf_cookie
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import NotFound, ValidationError
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

try:  # optional at import time; present via django-rest-knox
    from knox.models import AuthToken
except Exception:  # pragma: no cover
    AuthToken = None  # type: ignore

from . import models, selectors, serializers, services
try:  # avoid hard dependency at import time
    from apps.accounts import models as account_models
    from apps.accounts import services as account_services
except Exception:  # pragma: no cover
    account_models = None  # type: ignore
    account_services = None  # type: ignore


logger = logging.getLogger(__name__)


def _log_user_action(request, action):
    if not account_services or not account_models:
        return
    try:
        user = getattr(request, 'user', None)
        account_services.log_user_action(user=user, action=action)
    except Exception:
        pass


def _scenario_action_for_request(request, *, scenario_action, module_action):
    ref = (request.META.get('HTTP_REFERER') or '').lower()
    if '/data-management/test-modules/' in ref:
        return module_action
    if '/automation/test-scenarios/' in ref:
        return scenario_action
    # fallback: if the client explicitly sends a module id, treat as module-based
    try:
        data = getattr(request, 'data', None) or {}
        module_id = data.get('module') or data.get('module_id')
        if module_id not in (None, '', 0):
            return module_action
    except Exception:
        pass
    return scenario_action


def _automation_run_action(triggered_in: str | None):
    if not account_models:
        return None
    text = (triggered_in or '').lower()
    if 'project' in text:
        return account_models.UserAuditTrail.Actions.RUN_AUTOMATION_PROJECT
    if 'module' in text:
        return account_models.UserAuditTrail.Actions.RUN_AUTOMATION_MODULE
    if 'scenario' in text:
        return account_models.UserAuditTrail.Actions.RUN_AUTOMATION_SCENARIO
    if 'case' in text or 'testcase' in text:
        return account_models.UserAuditTrail.Actions.RUN_AUTOMATION_TEST_CASE
    return None

DEFAULT_AES_KEY = "kRdVzIqmQsfpRGItSLP5SDz0jkRLO9Cm"
DEFAULT_AES_IV = "1gJFNMeeQODA7wJA"
DEFAULT_CHANNEL_KEY = "dgzCF9eJw2uX9LNV4JrkQLxSHxBlZeGV"


def _get_paynamics_crypto_material() -> tuple[bytes | None, bytes | None, str | None]:
    key = getattr(settings, "PAYNAMICS_AES_KEY", "") or os.environ.get("PAYNAMICS_AES_KEY") or DEFAULT_AES_KEY
    iv = getattr(settings, "PAYNAMICS_AES_IV", "") or os.environ.get("PAYNAMICS_AES_IV") or DEFAULT_AES_IV
    channel_key = (
        getattr(settings, "PAYNAMICS_CHANNEL_KEY", "")
        or os.environ.get("PAYNAMICS_CHANNEL_KEY")
        or DEFAULT_CHANNEL_KEY
    )
    try:
        key_bytes = key.encode("utf-8") if key else None
        iv_bytes = iv.encode("utf-8") if iv else None
    except Exception:
        key_bytes = None
        iv_bytes = None
    return key_bytes, iv_bytes, channel_key


def _Crypto_available() -> bool:
    return bool(_AES and _pad and _unpad)


def _aes_encrypt_text(plaintext: str, key_bytes: bytes, iv_bytes: bytes) -> str:
    cipher = _AES.new(key_bytes, _AES.MODE_CBC, iv_bytes)
    padded = _pad(plaintext.encode("utf-8"), _AES.block_size)
    return base64.b64encode(cipher.encrypt(padded)).decode("utf-8")


def _aes_decrypt_text(encoded: str, key_bytes: bytes, iv_bytes: bytes) -> str:
    cipher = _AES.new(key_bytes, _AES.MODE_CBC, iv_bytes)
    decrypted = cipher.decrypt(base64.b64decode(encoded))
    return _unpad(decrypted, _AES.block_size).decode("utf-8")


def _compute_paynamics_signature(amount: str, pay_reference: str, pchannel: str, key_bytes: bytes, iv_bytes: bytes, channel_key: str | None) -> str:
    secret = channel_key or ""
    signature_input = f"{amount}{pay_reference}{pchannel}{secret}"
    return _aes_encrypt_text(signature_input, key_bytes, iv_bytes)


def _attempt_decrypt_response_data(encrypted_value: str) -> tuple[str | None, Any | None]:
    if not isinstance(encrypted_value, str) or not encrypted_value.strip():
        return None, None

    key_bytes, iv_bytes, _ = _get_paynamics_crypto_material()
    if not key_bytes or not iv_bytes:
        return None, None

    if not _Crypto_available():
        logger.warning("[tester.execute] PyCryptodome not available; unable to decrypt response data.")
        print("[tester.execute] decrypt skipped (missing PyCryptodome)")
        return None, None

    try:
        plaintext = _aes_decrypt_text(encrypted_value, key_bytes, iv_bytes)
    except Exception as exc:  # pragma: no cover - diagnostics path
        logger.exception("[tester.execute] failed to decrypt response data: %s", exc)
        print(f"[tester.execute] decrypt failure: {exc}")
        return None, None

    parsed: Any | None
    try:
        parsed = json.loads(plaintext)
    except json.JSONDecodeError:
        parsed = None

    return plaintext, parsed


def _apply_pay_reference_override(resolved_json: dict[str, Any], overrides: dict[str, Any]) -> tuple[str | None, dict[str, Any] | None, bool]:
    if not isinstance(resolved_json, dict) or not overrides:
        return None, None, False
    pay_reference = overrides.get("pay_reference") or overrides.get("dependency_value")
    if not pay_reference:
        return None, None, False
    data_field = resolved_json.get("data")
    if not isinstance(data_field, str) or not data_field.strip():
        return None, None, False

    key_bytes, iv_bytes, channel_key = _get_paynamics_crypto_material()
    if not key_bytes or not iv_bytes or not _Crypto_available():
        return None, None, False

    try:
        plaintext = _aes_decrypt_text(data_field, key_bytes, iv_bytes)
        payload = json.loads(plaintext)
    except Exception as exc:
        logger.exception("[tester.execute] failed to decode outbound payload: %s", exc)
        return None, None, False

    if not isinstance(payload, dict):
        return None, None, False

    amount = str(payload.get("amount") or "")
    pchannel = str(payload.get("pchannel") or "")
    pay_reference_str = str(pay_reference)
    changed = False

    if payload.get("pay_reference") != pay_reference_str:
        payload["pay_reference"] = pay_reference_str
        changed = True

    if amount and pchannel:
        try:
            payload["signature"] = _compute_paynamics_signature(
                amount,
                pay_reference_str,
                pchannel,
                key_bytes,
                iv_bytes,
                channel_key,
            )
            changed = True
        except Exception as exc:
            logger.exception("[tester.execute] failed to recompute signature: %s", exc)

    if not changed:
        return plaintext, payload, False

    updated_plaintext = json.dumps(payload, separators=(",", ":"))
    try:
        resolved_json["data"] = _aes_encrypt_text(updated_plaintext, key_bytes, iv_bytes)
    except Exception as exc:
        logger.exception("[tester.execute] failed to re-encrypt outbound payload: %s", exc)
        return plaintext, payload, False

    return updated_plaintext, payload, True


class ApiEnvironmentViewSet(viewsets.ModelViewSet):
    serializer_class = serializers.ApiEnvironmentSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return selectors.api_environment_list()

    def perform_create(self, serializer):
        super().perform_create(serializer)
        if account_models:
            _log_user_action(self.request, account_models.UserAuditTrail.Actions.CREATE_API_ENVIRONMENT)

    def perform_update(self, serializer):
        super().perform_update(serializer)
        if account_models:
            _log_user_action(self.request, account_models.UserAuditTrail.Actions.UPDATE_API_ENVIRONMENT)

    def perform_destroy(self, instance):
        super().perform_destroy(instance)
        if account_models:
            _log_user_action(self.request, account_models.UserAuditTrail.Actions.DELETE_API_ENVIRONMENT)


class ApiCollectionViewSet(viewsets.ModelViewSet):
    serializer_class = serializers.ApiCollectionSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return selectors.api_collection_list()

    def get_object(self):
        queryset = selectors.api_collection_base_queryset()
        try:
            return queryset.get(pk=self.kwargs["pk"])
        except models.ApiCollection.DoesNotExist as exc:  # pragma: no cover
            raise Http404 from exc

    @action(detail=True, methods=["post"], url_path="run")
    def run(self, request, pk=None):  # noqa: D401
        collection = self.get_object()
        environment = None
        environment_id = request.data.get("environment")
        if environment_id is not None:
            environment = models.ApiEnvironment.objects.filter(pk=environment_id).first()
            if environment is None:
                raise NotFound("Environment not found")

        overrides = request.data.get("overrides") or {}
        if not isinstance(overrides, dict):
            raise ValidationError({"overrides": "Overrides must be an object"})

        user: Any = request.user if request.user.is_authenticated else None
        run = services.run_collection(
            collection=collection,
            environment=environment,
            overrides=overrides,
            user=user,
        )
        serializer = serializers.ApiRunSerializer(run, context=self.get_serializer_context())
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=["post"], url_path="import-postman")
    def import_postman(self, request):
        file_obj = request.FILES.get("file")
        raw_payload: Any = None

        if file_obj is not None:
            try:
                raw_text = file_obj.read().decode("utf-8")
            except UnicodeDecodeError as exc:
                raise ValidationError({"file": "File must be UTF-8 encoded JSON."}) from exc
            try:
                raw_payload = json.loads(raw_text)
            except json.JSONDecodeError as exc:
                raise ValidationError({"file": "Invalid JSON file."}) from exc
        else:
            payload = request.data.get("collection")
            if isinstance(payload, (dict, list)):
                raw_payload = payload
            elif isinstance(payload, str) and payload.strip():
                try:
                    raw_payload = json.loads(payload)
                except json.JSONDecodeError as exc:
                    raise ValidationError({"collection": "Invalid JSON payload."}) from exc

        if raw_payload is None:
            raise ValidationError({"collection": "Postman collection JSON is required."})
        if not isinstance(raw_payload, dict):
            raise ValidationError({"collection": "Collection must be a JSON object."})

        try:
            collection = services.import_postman_collection(raw_payload)
        except ValueError as exc:
            raise ValidationError({"collection": str(exc)}) from exc

        serializer = self.get_serializer(collection)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class ApiRequestViewSet(viewsets.ModelViewSet):
    serializer_class = serializers.ApiRequestSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = models.ApiRequest.objects.select_related("collection").prefetch_related("assertions").order_by("order", "id")
        collection_id = self.request.query_params.get("collection")
        if collection_id:
            queryset = queryset.filter(collection_id=collection_id)
        return queryset

    @action(detail=True, methods=["get"], url_path="last-run")
    def last_run(self, request, pk=None):
        api_request = self.get_object()
        result = (
            models.ApiRunResult.objects.filter(request=api_request)
            .select_related("run__environment", "request")
            .order_by("-created_at")
            .first()
        )
        if result is None:
            return Response(None, status=status.HTTP_200_OK)
        serializer = serializers.ApiRunResultSerializer(result, context=self.get_serializer_context())
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=False, methods=["post"], url_path="reorder")
    def reorder(self, request):
        collection_id = request.data.get("collection")
        if collection_id in (None, ""):
            raise ValidationError({"collection": "Collection is required."})
        try:
            collection_id = int(collection_id)
        except (TypeError, ValueError) as exc:
            raise ValidationError({"collection": "Collection must be a valid integer."}) from exc

        directory_id = request.data.get("directory")
        if directory_id in (None, ""):
            directory_id = None
        else:
            try:
                directory_id = int(directory_id)
            except (TypeError, ValueError) as exc:
                raise ValidationError({"directory": "Directory must be a valid integer or null."}) from exc

        ordered_ids = request.data.get("ordered_ids") or []
        if not isinstance(ordered_ids, list):
            raise ValidationError({"ordered_ids": "ordered_ids must be a list."})
        try:
            ordered_ids = [int(item) for item in ordered_ids]
        except (TypeError, ValueError) as exc:
            raise ValidationError({"ordered_ids": "ordered_ids must contain only integers."}) from exc

        queryset = models.ApiRequest.objects.filter(collection_id=collection_id)
        if directory_id is None:
            queryset = queryset.filter(directory__isnull=True)
        else:
            queryset = queryset.filter(directory_id=directory_id)

        existing_ids = list(queryset.values_list("id", flat=True))
        if len(existing_ids) != len(ordered_ids) or set(existing_ids) != set(ordered_ids):
            raise ValidationError({"ordered_ids": "ordered_ids must match the existing request ids."})

        with transaction.atomic():
            for index, request_id in enumerate(ordered_ids):
                models.ApiRequest.objects.filter(pk=request_id).update(order=index)

        return Response({"status": "ok"}, status=status.HTTP_200_OK)
class ApiCollectionDirectoryViewSet(viewsets.ModelViewSet):
    serializer_class = serializers.ApiCollectionDirectorySerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = models.ApiCollectionDirectory.objects.select_related("collection", "parent").prefetch_related("requests").order_by("parent_id", "order", "id")
        collection_id = self.request.query_params.get("collection")
        if collection_id:
            queryset = queryset.filter(collection_id=collection_id)
        parent_id = self.request.query_params.get("parent")
        if parent_id:
            queryset = queryset.filter(parent_id=parent_id)
        return queryset

    def perform_create(self, serializer):
        collection = serializer.validated_data["collection"]
        parent = serializer.validated_data.get("parent")
        if "order" not in serializer.validated_data:
            sibling_count = models.ApiCollectionDirectory.objects.filter(collection=collection, parent=parent).count()
            serializer.save(order=sibling_count)
            return
        serializer.save()

    @action(detail=False, methods=["post"], url_path="reorder")
    def reorder(self, request):
        collection_id = request.data.get("collection")
        if collection_id in (None, ""):
            raise ValidationError({"collection": "Collection is required."})
        try:
            collection_id = int(collection_id)
        except (TypeError, ValueError) as exc:
            raise ValidationError({"collection": "Collection must be a valid integer."}) from exc

        parent_id = request.data.get("parent")
        if parent_id in (None, ""):
            parent_id = None
        else:
            try:
                parent_id = int(parent_id)
            except (TypeError, ValueError) as exc:
                raise ValidationError({"parent": "Parent must be a valid integer or null."}) from exc

        ordered_ids = request.data.get("ordered_ids") or []
        if not isinstance(ordered_ids, list):
            raise ValidationError({"ordered_ids": "ordered_ids must be a list."})
        try:
            ordered_ids = [int(item) for item in ordered_ids]
        except (TypeError, ValueError) as exc:
            raise ValidationError({"ordered_ids": "ordered_ids must contain only integers."}) from exc

        queryset = models.ApiCollectionDirectory.objects.filter(collection_id=collection_id)
        if parent_id is None:
            queryset = queryset.filter(parent__isnull=True)
        else:
            queryset = queryset.filter(parent_id=parent_id)

        existing_ids = list(queryset.values_list("id", flat=True))
        if len(existing_ids) != len(ordered_ids) or set(existing_ids) != set(ordered_ids):
            raise ValidationError({"ordered_ids": "ordered_ids must match the existing folder ids."})

        with transaction.atomic():
            for index, directory_id in enumerate(ordered_ids):
                models.ApiCollectionDirectory.objects.filter(pk=directory_id).update(order=index)

        return Response({"status": "ok"}, status=status.HTTP_200_OK)


class ApiRunViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = serializers.ApiRunSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return selectors.api_run_list()

    def get_object(self):
        instance = selectors.api_run_get(self.kwargs["pk"])
        if not instance:
            raise Http404
        return instance


class ProjectViewSet(viewsets.ModelViewSet):
    serializer_class = serializers.ProjectSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return selectors.project_list()

    def get_object(self):
        instance = selectors.project_get(self.kwargs["pk"])
        if instance is None:
            raise Http404
        return instance

    def perform_create(self, serializer):
        instance = serializer.save()
        if account_models:
            _log_user_action(self.request, account_models.UserAuditTrail.Actions.CREATE_PROJECT)
        return instance

    def perform_update(self, serializer):
        instance = serializer.save()
        if account_models:
            _log_user_action(self.request, account_models.UserAuditTrail.Actions.UPDATE_PROJECT)
        return instance


class TestScenarioViewSet(viewsets.ModelViewSet):
    serializer_class = serializers.TestScenarioSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = [JSONParser, FormParser, MultiPartParser]

    def get_queryset(self):
        queryset = selectors.test_scenario_list()
        project_param = self.request.query_params.get("project")
        if project_param in (None, ""):
            project_param = self.request.query_params.get("plan")
        if project_param not in (None, ""):
            try:
                queryset = queryset.filter(project_id=int(project_param))
            except (TypeError, ValueError) as exc:
                raise ValidationError({"project": "Project must be an integer."}) from exc
        module = self.request.query_params.get("module")
        if module:
            queryset = queryset.filter(module_id=module)
        return queryset

    def perform_create(self, serializer):
        instance = serializer.save()
        if account_models:
            action = _scenario_action_for_request(
                self.request,
                scenario_action=account_models.UserAuditTrail.Actions.CREATE_SCENARIO,
                module_action=account_models.UserAuditTrail.Actions.CREATE_SCENARIO_FROM_MODULE,
            )
            _log_user_action(self.request, action)
        return instance

    def perform_update(self, serializer):
        instance = serializer.save()
        if account_models:
            action = _scenario_action_for_request(
                self.request,
                scenario_action=account_models.UserAuditTrail.Actions.UPDATE_SCENARIO,
                module_action=account_models.UserAuditTrail.Actions.UPDATE_SCENARIO_FROM_MODULE,
            )
            _log_user_action(self.request, action)
        return instance

    def perform_destroy(self, instance):
        super().perform_destroy(instance)
        if account_models:
            action = _scenario_action_for_request(
                self.request,
                scenario_action=account_models.UserAuditTrail.Actions.DELETE_SCENARIO,
                module_action=account_models.UserAuditTrail.Actions.DELETE_SCENARIO_FROM_MODULE,
            )
            _log_user_action(self.request, action)

    def _is_allowed_scenario_attachment(self, filename: str | None, content_type: str | None) -> bool:
        """Allow images, videos, and common document formats."""

        import os

        name = (filename or "").lower()
        ext = ""
        try:
            _, ext = os.path.splitext(name)
        except Exception:
            ext = ""

        allowed_ext = {
            ".png", ".jpg", ".jpeg", ".gif", ".webp",
            ".mp4", ".webm", ".mov",
            ".csv",
            ".xls", ".xlsx",
            ".pdf",
            ".doc", ".docx",
        }
        if ext in allowed_ext:
            return True

        ct = (content_type or "").lower()
        if ct.startswith("image/") or ct.startswith("video/"):
            return True

        allowed_ct = {
            "text/csv",
            "application/csv",
            "application/pdf",
            "application/msword",
            "application/vnd.ms-excel",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        }
        return ct in allowed_ct

    @action(detail=True, methods=["get", "post"], url_path="attachments")
    def attachments(self, request, pk=None):
        scenario = self.get_object()

        if request.method == "GET":
            qs = models.TestScenarioAttachment.objects.filter(scenario=scenario).order_by("-created_at", "-id")
            return Response(
                serializers.TestScenarioAttachmentSerializer(qs, many=True, context={"request": request}).data
            )

        files = []
        try:
            if hasattr(request, "FILES"):
                files = request.FILES.getlist("files") or request.FILES.getlist("file")
        except Exception:
            files = []

        if not files:
            raise ValidationError({"files": "No files uploaded."})

        created: list[models.TestScenarioAttachment] = []
        for f in files:
            fname = getattr(f, "name", None) or "upload.bin"
            ctype = getattr(f, "content_type", None) or "application/octet-stream"
            if not self._is_allowed_scenario_attachment(fname, ctype):
                raise ValidationError({"files": f"File type not allowed: {fname}"})

            obj = models.TestScenarioAttachment(
                scenario=scenario,
                uploaded_by=request.user if getattr(request, "user", None) and request.user.is_authenticated else None,
                original_name=fname,
                content_type=ctype,
                size=getattr(f, "size", 0) or 0,
            )
            obj.file.save(fname, f, save=True)
            created.append(obj)

        qs = models.TestScenarioAttachment.objects.filter(id__in=[o.id for o in created]).order_by("-created_at", "-id")
        return Response(
            serializers.TestScenarioAttachmentSerializer(qs, many=True, context={"request": request}).data,
            status=status.HTTP_201_CREATED,
        )


class ScenarioCommentViewSet(viewsets.ModelViewSet):
    serializer_class = serializers.ScenarioCommentSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = [JSONParser, FormParser, MultiPartParser]

    def get_queryset(self):
        queryset = models.ScenarioComment.objects.select_related("user", "scenario").prefetch_related(
            "replies__user",
            "likes",
            "reactions",
            "replies__reactions",
            "attachments",
            "replies__attachments",
        ).all()
        scenario_id = self.request.query_params.get("scenario")
        if scenario_id:
            queryset = queryset.filter(scenario_id=scenario_id)
        # Only filter to top-level comments for list action, not for retrieve/update/delete
        if self.action == 'list':
            queryset = queryset.filter(parent__isnull=True)
        return queryset

    def _is_allowed_scenario_comment_attachment(self, filename: str | None, content_type: str | None) -> bool:
        """Allow images, videos, and common document formats."""

        import os

        name = (filename or "").lower()
        ext = ""
        try:
            _, ext = os.path.splitext(name)
        except Exception:
            ext = ""

        allowed_ext = {
            ".png", ".jpg", ".jpeg", ".gif", ".webp",
            ".mp4", ".webm", ".mov",
            ".csv",
            ".xls", ".xlsx",
            ".pdf",
            ".doc", ".docx",
        }
        if ext in allowed_ext:
            return True

        ct = (content_type or "").lower()
        if ct.startswith("image/") or ct.startswith("video/"):
            return True

        allowed_ct = {
            "text/csv",
            "application/csv",
            "application/pdf",
            "application/msword",
            "application/vnd.ms-excel",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        }
        return ct in allowed_ct

    @action(detail=True, methods=["get", "post"], url_path="attachments")
    def attachments(self, request, pk=None):
        comment = self.get_object()

        if request.method == "GET":
            qs = models.ScenarioCommentAttachment.objects.filter(comment=comment).order_by("-created_at", "-id")
            return Response(
                serializers.ScenarioCommentAttachmentSerializer(qs, many=True, context={"request": request}).data
            )

        files = []
        try:
            if hasattr(request, "FILES"):
                files = request.FILES.getlist("files") or request.FILES.getlist("file")
        except Exception:
            files = []

        if not files:
            raise ValidationError({"files": "No files uploaded."})

        created: list[models.ScenarioCommentAttachment] = []
        for f in files:
            fname = getattr(f, "name", None) or "upload.bin"
            ctype = getattr(f, "content_type", None) or "application/octet-stream"
            if not self._is_allowed_scenario_comment_attachment(fname, ctype):
                raise ValidationError({"files": f"File type not allowed: {fname}"})

            obj = models.ScenarioCommentAttachment(
                comment=comment,
                uploaded_by=request.user if getattr(request, "user", None) and request.user.is_authenticated else None,
                original_name=fname,
                content_type=ctype,
                size=getattr(f, "size", 0) or 0,
            )
            obj.file.save(fname, f, save=True)
            created.append(obj)

        qs = models.ScenarioCommentAttachment.objects.filter(id__in=[o.id for o in created]).order_by("-created_at", "-id")
        return Response(
            serializers.ScenarioCommentAttachmentSerializer(qs, many=True, context={"request": request}).data,
            status=status.HTTP_201_CREATED,
        )

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

    def perform_update(self, serializer):
        # Only allow users to edit their own comments
        if serializer.instance.user != self.request.user:
            raise ValidationError("You can only edit your own comments.")
        serializer.save()

    def perform_destroy(self, instance):
        # Only allow users to delete their own comments
        if instance.user != self.request.user:
            raise ValidationError("You can only delete your own comments.")
        instance.delete()

    @action(detail=True, methods=['post'])
    def toggle_like(self, request, pk=None):
        """Toggle like on a comment. Returns the updated like status."""
        comment = self.get_object()
        like, created = models.CommentLike.objects.get_or_create(
            comment=comment,
            user=request.user
        )
        
        if not created:
            # If like already existed, delete it (unlike)
            like.delete()
            liked = False
        else:
            liked = True

        # Return updated counts (compute fresh count; comment.likes may be prefetched/cached)
        likes_count = models.CommentLike.objects.filter(comment_id=comment.id).count()
        return Response({
            'liked': liked,
            'likes_count': likes_count,
        })

    @action(detail=True, methods=['post'])
    def set_reaction(self, request, pk=None):
        """Set or remove the current user's reaction on a comment."""
        comment = self.get_object()
        reaction = (request.data.get('reaction') or '').strip()
        if not reaction:
            raise ValidationError({'reaction': 'Reaction is required.'})
        if len(reaction) > 16:
            raise ValidationError({'reaction': 'Reaction is too long.'})

        obj, created = models.CommentReaction.objects.get_or_create(
            comment=comment,
            user=request.user,
            defaults={'reaction': reaction}
        )

        if not created:
            if obj.reaction == reaction:
                obj.delete()
                reacted = False
                current_reaction = None
            else:
                obj.reaction = reaction
                obj.save(update_fields=['reaction', 'updated_at'])
                reacted = True
                current_reaction = reaction
        else:
            reacted = True
            current_reaction = reaction

        reactions_summary = list(
            models.CommentReaction.objects.filter(comment_id=comment.id)
            .values('reaction')
            .annotate(count=Count('id'))
            .order_by('-count', 'reaction')
        )

        return Response({
            'reacted': reacted,
            'reaction': current_reaction,
            'reactions_summary': reactions_summary,
        })


class TestCaseCommentViewSet(viewsets.ModelViewSet):
    serializer_class = serializers.TestCaseCommentSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = [JSONParser, FormParser, MultiPartParser]

    def get_queryset(self):
        queryset = models.TestCaseComment.objects.select_related("user", "test_case").prefetch_related(
            "replies__user",
            "likes",
            "reactions",
            "replies__reactions",
            "attachments",
            "replies__attachments",
        ).all()
        test_case_id = self.request.query_params.get("test_case")
        if test_case_id:
            queryset = queryset.filter(test_case_id=test_case_id)
        if self.action == 'list':
            queryset = queryset.filter(parent__isnull=True)
        return queryset

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

    def perform_update(self, serializer):
        if serializer.instance.user != self.request.user:
            raise ValidationError("You can only edit your own comments.")
        serializer.save()

    def perform_destroy(self, instance):
        if instance.user != self.request.user:
            raise ValidationError("You can only delete your own comments.")
        instance.delete()

    def _is_allowed_comment_attachment(self, filename: str | None, content_type: str | None) -> bool:
        """Allow images, videos, and common document formats."""

        import os

        name = (filename or "").lower()
        ext = ""
        try:
            _, ext = os.path.splitext(name)
        except Exception:
            ext = ""

        allowed_ext = {
            ".png", ".jpg", ".jpeg", ".gif", ".webp",
            ".mp4", ".webm", ".mov",
            ".csv",
            ".xls", ".xlsx",
            ".pdf",
            ".doc", ".docx",
        }
        if ext in allowed_ext:
            return True

        ct = (content_type or "").lower()
        if ct.startswith("image/") or ct.startswith("video/"):
            return True

        allowed_ct = {
            "text/csv",
            "application/csv",
            "application/pdf",
            "application/msword",
            "application/vnd.ms-excel",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        }
        return ct in allowed_ct

    @action(detail=True, methods=["get", "post"], url_path="attachments")
    def attachments(self, request, pk=None):
        comment = self.get_object()

        if request.method == "GET":
            qs = models.TestCaseCommentAttachment.objects.filter(comment=comment).order_by("-created_at", "-id")
            return Response(
                serializers.TestCaseCommentAttachmentSerializer(qs, many=True, context={"request": request}).data
            )

        files = []
        try:
            if hasattr(request, "FILES"):
                files = request.FILES.getlist("files") or request.FILES.getlist("file")
        except Exception:
            files = []

        if not files:
            raise ValidationError({"files": "No files uploaded."})

        created: list[models.TestCaseCommentAttachment] = []
        for f in files:
            fname = getattr(f, "name", None) or "upload.bin"
            ctype = getattr(f, "content_type", None) or "application/octet-stream"
            if not self._is_allowed_comment_attachment(fname, ctype):
                raise ValidationError({"files": f"File type not allowed: {fname}"})

            obj = models.TestCaseCommentAttachment(
                comment=comment,
                uploaded_by=request.user if getattr(request, "user", None) and request.user.is_authenticated else None,
                original_name=fname,
                content_type=ctype,
                size=getattr(f, "size", 0) or 0,
            )
            obj.file.save(fname, f, save=True)
            created.append(obj)

        qs = models.TestCaseCommentAttachment.objects.filter(id__in=[o.id for o in created]).order_by("-created_at", "-id")
        return Response(
            serializers.TestCaseCommentAttachmentSerializer(qs, many=True, context={"request": request}).data,
            status=status.HTTP_201_CREATED,
        )

    @action(detail=True, methods=['post'])
    def toggle_like(self, request, pk=None):
        comment = self.get_object()
        like, created = models.TestCaseCommentLike.objects.get_or_create(
            comment=comment,
            user=request.user
        )

        if not created:
            like.delete()
            liked = False
        else:
            liked = True

        likes_count = models.TestCaseCommentLike.objects.filter(comment_id=comment.id).count()
        return Response({
            'liked': liked,
            'likes_count': likes_count,
        })

    @action(detail=True, methods=['post'])
    def set_reaction(self, request, pk=None):
        comment = self.get_object()
        reaction = (request.data.get('reaction') or '').strip()
        if not reaction:
            raise ValidationError({'reaction': 'Reaction is required.'})
        if len(reaction) > 16:
            raise ValidationError({'reaction': 'Reaction is too long.'})

        obj, created = models.TestCaseCommentReaction.objects.get_or_create(
            comment=comment,
            user=request.user,
            defaults={'reaction': reaction}
        )

        if not created:
            if obj.reaction == reaction:
                obj.delete()
                reacted = False
                current_reaction = None
            else:
                obj.reaction = reaction
                obj.save(update_fields=['reaction', 'updated_at'])
                reacted = True
                current_reaction = reaction
        else:
            reacted = True
            current_reaction = reaction

        reactions_summary = list(
            models.TestCaseCommentReaction.objects.filter(comment_id=comment.id)
            .values('reaction')
            .annotate(count=Count('id'))
            .order_by('-count', 'reaction')
        )

        return Response({
            'reacted': reacted,
            'reaction': current_reaction,
            'reactions_summary': reactions_summary,
        })


class TestCaseViewSet(viewsets.ModelViewSet):
    serializer_class = serializers.TestCaseSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = [JSONParser, FormParser, MultiPartParser]

    def update(self, request, *args, **kwargs):
        try:
            user = getattr(request, 'user', None)
            user_repr = getattr(user, 'email', None) or getattr(user, 'username', None) or str(user)
            ip = request.META.get('HTTP_X_FORWARDED_FOR', request.META.get('REMOTE_ADDR'))
            logger.debug('[core] TestCaseViewSet.update incoming data: user=%s ip=%s data=%s', user_repr, ip, getattr(request, 'data', None))
        except Exception:
            pass
        # Defensive: prevent clients from changing the testcase_id on update
        try:
            instance = self.get_object()
            incoming = getattr(request, 'data', {}) or {}
            if 'testcase_id' in incoming and incoming.get('testcase_id') not in (None, '', str(instance.testcase_id)):
                raise ValidationError({'testcase_id': 'testcase_id cannot be changed once created.'})
        except ValidationError:
            raise
        except Exception:
            # ignore failures to inspect instance; proceed to allow serializer to handle
            pass
        response = super().update(request, *args, **kwargs)
        if account_models:
            _log_user_action(request, account_models.UserAuditTrail.Actions.UPDATE_TEST_CASE)
        # Post-update: ensure owner is set when an authenticated user made
        # the update but the instance has no owner (client may have sent
        # an empty owner value). We do this after the serializer has saved
        # to avoid interfering with normal update validation and behavior.
        try:
            inst = self.get_object()
            if (
                inst is not None
                and getattr(request, 'user', None)
                and request.user.is_authenticated
                and getattr(inst, 'owner', None) is None
            ):
                inst.owner = request.user
                inst.save(update_fields=['owner'])
        except Exception:
            # Don't let fallback failures break the response
            pass
        return response

    def partial_update(self, request, *args, **kwargs):
        try:
            user = getattr(request, 'user', None)
            user_repr = getattr(user, 'email', None) or getattr(user, 'username', None) or str(user)
            ip = request.META.get('HTTP_X_FORWARDED_FOR', request.META.get('REMOTE_ADDR'))
            logger.debug('[core] TestCaseViewSet.partial_update incoming data: user=%s ip=%s data=%s', user_repr, ip, getattr(request, 'data', None))
        except Exception:
            pass
        # Defensive: prevent clients from changing the testcase_id on partial update
        try:
            instance = self.get_object()
            incoming = getattr(request, 'data', {}) or {}
            if 'testcase_id' in incoming and incoming.get('testcase_id') not in (None, '', str(instance.testcase_id)):
                raise ValidationError({'testcase_id': 'testcase_id cannot be changed once created.'})
        except ValidationError:
            raise
        except Exception:
            pass
        response = super().partial_update(request, *args, **kwargs)
        if account_models:
            _log_user_action(request, account_models.UserAuditTrail.Actions.UPDATE_TEST_CASE)
        # same post-update owner assignment for partial updates
        try:
            inst = self.get_object()
            if (
                inst is not None
                and getattr(request, 'user', None)
                and request.user.is_authenticated
                and getattr(inst, 'owner', None) is None
            ):
                inst.owner = request.user
                inst.save(update_fields=['owner'])
        except Exception:
            pass
        return response

    def destroy(self, request, *args, **kwargs):
        response = super().destroy(request, *args, **kwargs)
        if account_models:
            _log_user_action(request, account_models.UserAuditTrail.Actions.DELETE_TEST_CASE)
        return response

    def get_queryset(self):
        queryset = selectors.test_case_list()
        search = self.request.query_params.get("search")
        if search:
            # allow searching by testcase_id, title, description, precondition, requirements
            q = (
                Q(testcase_id__icontains=search)
                | Q(title__icontains=search)
                | Q(description__icontains=search)
                | Q(precondition__icontains=search)
                | Q(requirements__icontains=search)
            )
            # if the search looks like an integer, allow searching by primary key too
            try:
                if str(search).isdigit():
                    q |= Q(pk=int(search))
            except Exception:
                pass
            queryset = queryset.filter(q)

        project_param = self.request.query_params.get("project")
        if project_param in (None, ""):
            project_param = self.request.query_params.get("plan")
        if project_param not in (None, ""):
            try:
                queryset = queryset.filter(scenario__project_id=int(project_param))
            except (TypeError, ValueError) as exc:
                raise ValidationError({"project": "Project must be an integer."}) from exc

        scenario_id = self.request.query_params.get("scenario")
        if scenario_id:
            try:
                queryset = queryset.filter(scenario_id=int(scenario_id))
            except (TypeError, ValueError) as exc:
                raise ValidationError({"scenario": "Scenario must be an integer."}) from exc

        return queryset

    def _is_allowed_testcase_attachment(self, filename: str | None, content_type: str | None) -> bool:
        """Allow images, videos, and common document formats."""

        name = (filename or "").lower()
        ext = ""
        try:
            _, ext = os.path.splitext(name)
        except Exception:
            ext = ""

        allowed_ext = {
            ".png", ".jpg", ".jpeg", ".gif", ".webp",
            ".mp4", ".webm", ".mov",
            ".csv",
            ".xls", ".xlsx",
            ".pdf",
            ".doc", ".docx",
        }
        if ext in allowed_ext:
            return True

        ct = (content_type or "").lower()
        if ct.startswith("image/") or ct.startswith("video/"):
            return True

        allowed_ct = {
            "text/csv",
            "application/csv",
            "application/pdf",
            "application/msword",
            "application/vnd.ms-excel",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        }
        return ct in allowed_ct

    @action(detail=True, methods=["get", "post"], url_path="attachments")
    def attachments(self, request, pk=None):
        test_case = self.get_object()

        if request.method == "GET":
            qs = models.TestCaseAttachment.objects.filter(test_case=test_case).order_by("-created_at", "-id")
            return Response(serializers.TestCaseAttachmentSerializer(qs, many=True, context={"request": request}).data)

        # POST: accept multipart file upload(s)
        files = []
        try:
            if hasattr(request, "FILES"):
                files = request.FILES.getlist("files") or request.FILES.getlist("file")
        except Exception:
            files = []

        if not files:
            raise ValidationError({"files": "No files uploaded."})

        created: list[models.TestCaseAttachment] = []
        for f in files:
            try:
                fname = getattr(f, "name", None) or "upload.bin"
                ctype = getattr(f, "content_type", None) or "application/octet-stream"
                if not self._is_allowed_testcase_attachment(fname, ctype):
                    raise ValidationError({"files": f"Unsupported file type: {fname}"})

                obj = models.TestCaseAttachment.objects.create(
                    test_case=test_case,
                    file=f,
                    original_name=fname,
                    content_type=ctype,
                    size=int(getattr(f, "size", 0) or 0),
                    uploaded_by=getattr(request, "user", None) if getattr(request, "user", None) and request.user.is_authenticated else None,
                )
                created.append(obj)
            except ValidationError:
                raise
            except Exception as exc:
                logger.exception("[core] failed to save testcase attachment: %s", exc)
                raise ValidationError({"files": "Failed to upload attachment."})

        # Return refreshed list so the UI can re-render.
        qs = models.TestCaseAttachment.objects.filter(test_case=test_case).order_by("-created_at", "-id")
        return Response(serializers.TestCaseAttachmentSerializer(qs, many=True, context={"request": request}).data, status=status.HTTP_201_CREATED)

    def create(self, request, *args, **kwargs):
        # Defensive: some clients or parsers may omit the 'testcase_id' key
        # entirely which can cause certain validators or upstream code to
        # treat the field as required. Ensure the incoming data contains the
        # key set to None so the serializer can accept it and the model's
        # save() will auto-generate the id when appropriate.
        try:
            logger.debug('[core] TestCaseViewSet.create incoming data: %s', getattr(request, 'data', None))
        except Exception:
            pass
        data = request.data.copy() if hasattr(request.data, 'copy') else dict(request.data)
        if 'testcase_id' not in data or data.get('testcase_id') in (None, ''):
            data['testcase_id'] = None
        # If owner not provided, set to the authenticated user
        try:
            if not data.get('owner') and getattr(request, 'user', None) and request.user.is_authenticated:
                data['owner'] = request.user.id
        except Exception:
            pass
        serializer = self.get_serializer(data=data)
        try:
            serializer.is_valid(raise_exception=True)
        except ValidationError as exc:
            # Defensive: if the only error is that testcase_id is required,
            # inject a None value and retry so the model can auto-generate it.
            errors = getattr(exc, 'detail', {}) or {}
            if 'testcase_id' in errors and errors.get('testcase_id') in (["This field is required."],) or (
                isinstance(errors.get('testcase_id'), list) and any(str(e).lower().startswith('this field is required') for e in errors.get('testcase_id'))
            ):
                data['testcase_id'] = None
                serializer = self.get_serializer(data=data)
                serializer.is_valid(raise_exception=True)
            else:
                raise
        

        # Save using perform_create but ensure owner is set to authenticated user
        try:
            if getattr(request, 'user', None) and request.user.is_authenticated:
                serializer.save(owner=request.user)
            else:
                self.perform_create(serializer)
        except TypeError:
            # Fallback if serializer.save doesn't accept owner (older behavior)
            self.perform_create(serializer)
        # Ensure owner is set on the saved instance when the client did not
        # provide an owner id (or provided an empty value). Some clients may
        # send an empty string for 'owner' which should be treated as absent.
        try:
            inst = getattr(serializer, 'instance', None)
            if (
                inst is not None
                and getattr(request, 'user', None)
                and request.user.is_authenticated
                and getattr(inst, 'owner', None) is None
            ):
                inst.owner = request.user
                # save only owner field to avoid touching other fields
                inst.save(update_fields=['owner'])
        except Exception:
            # don't let a logging/save failure break the response
            pass
        # Defensive: ensure related_api_request from the incoming payload is
        # persisted on the instance. Some clients or serializer behaviors may
        # omit saving this FK during the initial create step; explicitly set
        # it here if provided in the request data.
        try:
            inst = getattr(serializer, 'instance', None)
            if inst is not None:
                # Prefer the local `data` dict (copied/normalized above) but
                # fall back to request.data if needed.
                incoming_related = None
                try:
                    incoming_related = data.get('related_api_request') if isinstance(data, dict) else None
                except Exception:
                    incoming_related = None
                if incoming_related is None and hasattr(request, 'data'):
                    incoming_related = getattr(request.data, 'get', lambda k, d=None: d)('related_api_request', None)
                # Normalize to integer id or None
                try:
                    if incoming_related in ('', None):
                        rid = None
                    else:
                        rid = int(incoming_related)
                except Exception:
                    rid = None
                if rid and getattr(inst, 'related_api_request_id', None) != rid:
                    try:
                        inst.related_api_request_id = rid
                        inst.save(update_fields=['related_api_request'])
                    except Exception:
                        # ignore failures here to avoid breaking the response
                        pass
                elif rid is None and getattr(inst, 'related_api_request_id', None) is not None:
                    try:
                        inst.related_api_request = None
                        inst.save(update_fields=['related_api_request'])
                    except Exception:
                        pass
        except Exception:
            # never fail the request response because of this best-effort step
            pass
        headers = self.get_success_headers(serializer.data)
        if account_models:
            _log_user_action(request, account_models.UserAuditTrail.Actions.CREATE_TEST_CASE)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)


class TestModulesViewSet(viewsets.ModelViewSet):
    serializer_class = serializers.TestModulesSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = models.TestModules.objects.order_by("title", "id")
        project_param = self.request.query_params.get("project")
        if project_param in (None, ""):
            project_param = self.request.query_params.get("plan")
        if project_param not in (None, ""):
            try:
                queryset = queryset.filter(project_id=int(project_param))
            except (TypeError, ValueError) as exc:
                raise ValidationError({"project": "Project must be an integer."}) from exc
        return queryset

    def perform_create(self, serializer):
        instance = serializer.save()
        if account_models:
            _log_user_action(self.request, account_models.UserAuditTrail.Actions.CREATE_MODULE)
        return instance

    def perform_update(self, serializer):
        instance = serializer.save()
        if account_models:
            _log_user_action(self.request, account_models.UserAuditTrail.Actions.UPDATE_MODULE)
        return instance

    def perform_destroy(self, instance):
        super().perform_destroy(instance)
        if account_models:
            _log_user_action(self.request, account_models.UserAuditTrail.Actions.DELETE_MODULE)


def _prepare_automation_data(*, automated_scenarios_only: bool = False) -> dict[str, Any]:
    projects_qs = selectors.project_list(automated_scenarios_only=automated_scenarios_only)
    projects_payload = serializers.ProjectSerializer(projects_qs, many=True).data
    environments_qs = selectors.api_environment_list()
    environments_payload = serializers.ApiEnvironmentSerializer(environments_qs, many=True).data
    test_modules_qs = models.TestModules.objects.order_by("title", "id")
    test_modules_payload = serializers.TestModulesSerializer(test_modules_qs, many=True).data

    scenario_count = sum(len(project.get("scenarios", [])) for project in projects_payload)
    case_count = sum(
        len(scenario.get("cases", []))
        for project in projects_payload
        for scenario in project.get("scenarios", [])
    )

    metrics = {
        "plans": len(projects_payload),
        "projects": len(projects_payload),
        "scenarios": scenario_count,
        "cases": case_count,
        "modules": len(test_modules_payload),
        "collections": models.ApiCollection.objects.count(),
        "runs": models.ApiRun.objects.count(),
        "environments": len(environments_payload),
        "risks": 0,
        "mitigation_plans": 0,
        "risk_mitigations": 0,
    }

    recent_runs = (
        models.ApiRun.objects.select_related("collection", "environment", "triggered_by")
        .order_by("-created_at")[:5]
    )
    highlighted_collections = models.ApiCollection.objects.order_by("name")[:6]

    api_endpoints = {
        "plans": reverse("core:core-test-plans-list"),
        "maintenances": "",
        "scenarios": reverse("core:core-test-scenarios-list"),
        "cases": reverse("core:core-test-cases-list"),
        "scopes": "",
        "collections": reverse("core:core-collections-list"),
        "environments": reverse("core:core-environments-list"),
        "runs": reverse("core:core-runs-list"),
        "risks": "",
        "mitigation_plans": "",
        "risk_mitigations": "",
        "test_tools": "",
        "test_modules": reverse("core:core-test-modules-list"),
        # non-router endpoints
        "tester_execute": reverse("core:core-request-execute"),
        "automation_report_finalize": reverse("core:core-automation-report-finalize"),
        "automation_report_create": reverse("core:core-automation-report-create"),
        "load_tests": reverse("core:core-load-tests"),
    }

    selected_plan = projects_payload[0] if projects_payload else None
    selected_scenario = None
    if selected_plan:
        scenarios = selected_plan.get("scenarios", [])
        if scenarios:
            selected_scenario = scenarios[0]

    return {
        "plans": projects_payload,
        "metrics": metrics,
        "recent_runs": recent_runs,
        "highlighted_collections": highlighted_collections,
        "api_endpoints": api_endpoints,
        "selected_plan": selected_plan,
        "selected_scenario": selected_scenario,
        "environments": environments_payload,
        "risks": [],
        "mitigation_plans": [],
        "risk_mitigations": [],
        "test_tools": [],
        "test_modules": test_modules_payload,
    }


@login_required
def automation_overview(request):
    """Render top-level Automation overview content."""

    data = _prepare_automation_data()
    context = {
        "initial_metrics": data["metrics"],
        "recent_runs": data["recent_runs"],
        "highlighted_collections": data["highlighted_collections"],
    }
    return render(request, "core/automation_overview.html", context)


@login_required
def automation_run(request):
    """Render the Automation run workspace with initial hierarchy data."""

    data = _prepare_automation_data(automated_scenarios_only=True)
    context = {
        "initial_projects": data["plans"],
        "api_endpoints": data["api_endpoints"],
        "initial_environments": data["environments"],
    }
    return render(request, "core/automation_run.html", context)


@login_required
def automation_load_testing(request):
    """Render the Automation load testing workspace (Locust-based)."""

    data = _prepare_automation_data(automated_scenarios_only=True)
    context = {
        "initial_projects": data["plans"],
        "api_endpoints": data["api_endpoints"],
    }
    return render(request, "core/automation_load_testing.html", context)



@login_required
def automation_reports(request):
    """Render a simple Automation reports page (placeholder).

    This page currently provides a landing area for future reports and basic
    metrics context so frontend code can attach charts or tables.
    """

    data = _prepare_automation_data()
    # Provide recent AutomationReport and ApiRunResultReport rows for server-side rendering
    try:
        automation_reports_qs = models.AutomationReport.objects.order_by("-created_at")[:50]
        automation_reports = serializers.AutomationReportSerializer(automation_reports_qs, many=True).data
    except Exception:
        automation_reports = []
    # Format started/finished dates for display in the template (YYYY-MM-DD hh:mm AM/PM)
    try:
        for ser, obj in zip(automation_reports, automation_reports_qs):
            try:
                if getattr(obj, 'started', None):
                    ser['started'] = obj.started.strftime('%Y-%m-%d %I:%M %p')
                else:
                    ser['started'] = '—'
            except Exception:
                ser['started'] = ser.get('started') or '—'
            try:
                if getattr(obj, 'finished', None):
                    ser['finished'] = obj.finished.strftime('%Y-%m-%d %I:%M %p')
                else:
                    ser['finished'] = '—'
            except Exception:
                ser['finished'] = ser.get('finished') or '—'
    except Exception:
        pass
    try:
        testcase_reports_qs = models.ApiRunResultReport.objects.select_related("testcase", "run", "request").order_by("-created_at")[:100]
        # base serialized list (from serializer)
        testcase_reports_serialized = serializers.ApiRunResultReportSerializer(testcase_reports_qs, many=True).data
        # Build a lookup of automation_report.id -> triggered_by (from serialized automation_reports)
        ar_triggered_by = {}
        try:
            for r in automation_reports:
                if isinstance(r, dict) and r.get('id') is not None:
                    ar_triggered_by[r['id']] = r.get('triggered_by')
        except Exception:
            ar_triggered_by = {}

        # enrich serialized items with automation_report id from the queryset objects and copy triggered_by
        testcase_reports = []
        for ser, obj in zip(testcase_reports_serialized, testcase_reports_qs):
            item = dict(ser)
            try:
                ar_id = getattr(obj, "automation_report_id", None)
            except Exception:
                ar_id = None
            item["automation_report_id"] = ar_id
            # prefer the serialized automation report's triggered_by value when available
            try:
                item["triggered_by"] = ar_triggered_by.get(ar_id) if ar_id is not None else None
            except Exception:
                item["triggered_by"] = None
            # attach nicely formatted started/finished timestamps for display
            try:
                if getattr(obj, 'created_at', None):
                    item['started'] = obj.created_at.strftime('%Y-%m-%d %I:%M %p')
                else:
                    item['started'] = ser.get('created_at') or '—'
            except Exception:
                item['started'] = ser.get('created_at') or '—'
            try:
                if getattr(obj, 'updated_at', None):
                    item['finished'] = obj.updated_at.strftime('%Y-%m-%d %I:%M %p')
                else:
                    item['finished'] = ser.get('updated_at') or '—'
            except Exception:
                item['finished'] = ser.get('updated_at') or '—'

            # Human-friendly outcome label for UI tables.
            # The underlying result-report status is one of: passed|failed|error.
            # Map error to Failed (UI expects Passed/Failed/Queued/Running).
            try:
                status_val = (getattr(obj, 'status', None) or item.get('status') or '')
                status_norm = str(status_val).strip().lower()
            except Exception:
                status_norm = ''
            outcome = ''
            if status_norm == 'passed':
                outcome = 'Passed'
            elif status_norm in ('failed', 'error'):
                outcome = 'Failed'
            else:
                # Fallback to run status when status is missing/unexpected.
                try:
                    run_status = str(getattr(getattr(obj, 'run', None), 'status', '') or '').lower()
                except Exception:
                    run_status = ''
                if run_status == 'running':
                    outcome = 'Running'
                elif run_status in ('pending', 'queued'):
                    outcome = 'Queued'
                else:
                    outcome = 'Queued'
            item['outcome'] = outcome

            # Back-compat keys for templates that reference created_at/finished_at.
            item['created_label'] = item.get('started') or '—'
            item['finished_at'] = item.get('finished') or '—'

            testcase_reports.append(item)
    except Exception:
        testcase_reports = []

    # For the parent AutomationReport table, compute "Finished At" from the
    # latest testcase run in that report (max ApiRunResultReport.updated_at).
    # This better reflects real completion time across different trigger paths.
    try:
        report_ids = [r.get("id") for r in automation_reports if isinstance(r, dict) and r.get("id") is not None]
        report_ids = [int(rid) for rid in report_ids if str(rid).isdigit()]
        if report_ids:
            last_finished_map = {}
            try:
                rows = (
                    models.ApiRunResultReport.objects.filter(automation_report_id__in=report_ids)
                    .values("automation_report_id")
                    .annotate(last_finished=Max("updated_at"))
                )
                for row in rows:
                    rid = row.get("automation_report_id")
                    dt = row.get("last_finished")
                    if rid is not None and dt is not None:
                        last_finished_map[int(rid)] = dt
            except Exception:
                last_finished_map = {}

            for r in automation_reports:
                try:
                    rid = r.get("id") if isinstance(r, dict) else None
                    if rid is None:
                        continue
                    dt = last_finished_map.get(int(rid))
                    if dt is not None:
                        r["finished"] = dt.strftime('%Y-%m-%d %I:%M %p')
                    else:
                        # Keep existing formatted value, but ensure it's not blank.
                        r["finished"] = r.get("finished") or '—'
                except Exception:
                    # Don't let formatting errors break the page.
                    try:
                        if isinstance(r, dict) and not r.get("finished"):
                            r["finished"] = '—'
                    except Exception:
                        pass
    except Exception:
        pass

    # attach per-report testcase lists so template can render "No test cases" correctly
    try:
        for r in automation_reports:
            rid = r.get("id")
            if rid is None:
                r["testcases"] = []
            else:
                r["testcases"] = [t for t in testcase_reports if t.get("automation_report_id") == rid]
    except Exception:
        # fall back to empty lists on any error
        for r in automation_reports:
            r["testcases"] = []

    context = {
        "initial_metrics": data.get("metrics", {}),
        "recent_runs": data.get("recent_runs", []),
        "automation_reports": automation_reports,
        # provide the enriched testcase reports for client-side use (JSON embed)
        "testcase_reports_serialized": testcase_reports,
    }
    return render(request, "core/automation_reports.html", context)


@login_required
def automation_ui_testing(request):
    """Render the Automation UI testing page (Playwright-based)."""

    context = {}
    return render(request, "core/automation_ui_testing.html", context)


@login_required
def automation_reports_export(request):
    """Export Automated Reports to an Excel file.

    Query params:
      - start: YYYY-MM-DD (optional)
      - end: YYYY-MM-DD (optional)

    Exports 3 sheets:
      - Summary: per-day totals and chart
      - Reports: parent AutomationReport rows
      - Testcases: child ApiRunResultReport rows
    """

    if account_models:
        _log_user_action(request, account_models.UserAuditTrail.Actions.EXPORT_AUTOMATED_REPORT)

    # Parse date range (inclusive)
    start_str = (request.GET.get("start") or request.GET.get("from") or "").strip()
    end_str = (request.GET.get("end") or request.GET.get("to") or "").strip()
    start_date = None
    end_date = None
    try:
        if start_str:
            start_date = timezone.datetime.fromisoformat(start_str).date()
    except Exception:
        start_date = None
    try:
        if end_str:
            end_date = timezone.datetime.fromisoformat(end_str).date()
    except Exception:
        end_date = None

    qs = models.AutomationReport.objects.select_related("triggered_by").order_by("-created_at")
    # Filter by report.started when present, else by created_at
    if start_date:
        qs = qs.filter(Q(started__date__gte=start_date) | Q(started__isnull=True, created_at__date__gte=start_date))
    if end_date:
        qs = qs.filter(Q(started__date__lte=end_date) | Q(started__isnull=True, created_at__date__lte=end_date))

    reports = list(qs)
    report_ids = [r.id for r in reports]

    # Compute "finished at" per report from latest testcase run (max updated_at)
    last_finished_map: dict[int, timezone.datetime] = {}
    if report_ids:
        try:
            rows = (
                models.ApiRunResultReport.objects.filter(automation_report_id__in=report_ids)
                .values("automation_report_id")
                .annotate(last_finished=Max("updated_at"))
            )
            for row in rows:
                rid = row.get("automation_report_id")
                dt = row.get("last_finished")
                if rid is not None and dt is not None:
                    last_finished_map[int(rid)] = dt
        except Exception:
            last_finished_map = {}

    # Load testcase reports
    testcase_rows = []
    if report_ids:
        try:
            testcase_qs = (
                models.ApiRunResultReport.objects.select_related("run", "request", "testcase")
                .filter(automation_report_id__in=report_ids)
                .order_by("automation_report_id", "order", "id")
            )
            for obj in testcase_qs:
                try:
                    status_norm = str(getattr(obj, "status", "") or "").strip().lower()
                except Exception:
                    status_norm = ""
                if status_norm == "passed":
                    outcome = "Passed"
                elif status_norm in ("failed", "error"):
                    outcome = "Failed"
                else:
                    # Fallback: if run is still in progress, show Running; otherwise Queued.
                    try:
                        run_status = str(getattr(getattr(obj, "run", None), "status", "") or "").lower()
                    except Exception:
                        run_status = ""
                    outcome = "Running" if run_status == "running" else "Queued"

                testcase_rows.append(
                    {
                        "automation_report_id": obj.automation_report_id,
                        "testcase_id": (obj.testcase.testcase_id if obj.testcase else None),
                        "request_name": (obj.request.name if obj.request else None),
                        "run_id": (obj.run_id or None),
                        "outcome": outcome,
                        "status": getattr(obj, "status", None),
                        "started_at": getattr(obj, "created_at", None),
                        "finished_at": getattr(obj, "updated_at", None),
                    }
                )
        except Exception:
            testcase_rows = []

    # Build summary by date (report-level), using report finished date derived from testcase rows.
    summary_by_day: dict[str, dict[str, int]] = {}
    for r in reports:
        dt = last_finished_map.get(r.id) or getattr(r, "finished", None) or getattr(r, "updated_at", None)
        day_key = None
        try:
            if dt:
                day_key = dt.date().isoformat()
        except Exception:
            day_key = None
        if not day_key:
            # Put unfinished reports into a synthetic bucket so user can still see them.
            day_key = "Unfinished"

        bucket = summary_by_day.setdefault(
            day_key,
            {"reports": 0, "passed": 0, "failed": 0, "blocked": 0, "total": 0},
        )
        bucket["reports"] += 1
        bucket["passed"] += int(getattr(r, "total_passed", 0) or 0)
        bucket["failed"] += int(getattr(r, "total_failed", 0) or 0)
        bucket["blocked"] += int(getattr(r, "total_blocked", 0) or 0)
        bucket["total"] += (
            int(getattr(r, "total_passed", 0) or 0)
            + int(getattr(r, "total_failed", 0) or 0)
            + int(getattr(r, "total_blocked", 0) or 0)
        )

    # Create workbook
    try:
        from openpyxl import Workbook
        from openpyxl.chart import LineChart, Reference
        from openpyxl.styles import Alignment, Font
    except Exception as exc:
        return HttpResponse(f"Excel export dependency not available: {exc}", status=500)

    wb = Workbook()
    # Remove default sheet
    try:
        wb.remove(wb.active)
    except Exception:
        pass

    ws_summary = wb.create_sheet("Summary")
    ws_reports = wb.create_sheet("Reports")
    ws_testcases = wb.create_sheet("Testcases")

    bold = Font(bold=True)

    # --- Summary sheet ---
    exported_at = timezone.now()
    range_label = f"{start_str or '—'} to {end_str or '—'}"
    ws_summary["A1"].value = "Automated Report Export"
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary["A2"].value = "Date range"
    ws_summary["A2"].font = bold
    ws_summary["B2"].value = range_label
    ws_summary["A3"].value = "Exported at"
    ws_summary["A3"].font = bold
    ws_summary["B3"].value = exported_at.strftime("%Y-%m-%d %H:%M")

    header_row = 5
    summary_headers = ["Day", "Reports", "Passed", "Failed", "Blocked", "Total"]
    for col, h in enumerate(summary_headers, start=1):
        cell = ws_summary.cell(row=header_row, column=col, value=h)
        cell.font = bold
        cell.alignment = Alignment(horizontal="center")

    # Sorted days (dates first), then Unfinished
    day_keys = sorted([k for k in summary_by_day.keys() if k != "Unfinished"])
    if "Unfinished" in summary_by_day:
        day_keys.append("Unfinished")

    r0 = header_row + 1
    for i, day in enumerate(day_keys):
        b = summary_by_day[day]
        ws_summary.cell(row=r0 + i, column=1, value=day)
        ws_summary.cell(row=r0 + i, column=2, value=b["reports"])
        ws_summary.cell(row=r0 + i, column=3, value=b["passed"])
        ws_summary.cell(row=r0 + i, column=4, value=b["failed"])
        ws_summary.cell(row=r0 + i, column=5, value=b["blocked"])
        ws_summary.cell(row=r0 + i, column=6, value=b["total"])

    ws_summary.freeze_panes = ws_summary["A6"]

    # Add chart (skip if no data rows)
    if day_keys:
        chart = LineChart()
        chart.title = "Totals by Day"
        chart.y_axis.title = "Count"
        chart.x_axis.title = "Day"
        data = Reference(ws_summary, min_col=3, max_col=5, min_row=header_row, max_row=r0 + len(day_keys) - 1)
        cats = Reference(ws_summary, min_col=1, min_row=r0, max_row=r0 + len(day_keys) - 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        # Ensure something is visible even for a single-day range.
        # With one category Excel has no line segment; markers make the point visible.
        for s in chart.series:
            try:
                s.marker.symbol = "circle"
                s.marker.size = 7
            except Exception:
                pass
            try:
                s.graphicalProperties.line.width = 20000
            except Exception:
                pass

        ws_summary.add_chart(chart, "H5")

    ws_summary.column_dimensions["A"].width = 14
    ws_summary.column_dimensions["B"].width = 10
    ws_summary.column_dimensions["C"].width = 10
    ws_summary.column_dimensions["D"].width = 10
    ws_summary.column_dimensions["E"].width = 10
    ws_summary.column_dimensions["F"].width = 10
    ws_summary.column_dimensions["H"].width = 18

    # --- Reports sheet ---
    report_headers = [
        "Report ID",
        "Triggered In",
        "Triggered By",
        "Passed",
        "Failed",
        "Blocked",
        "Started At",
        "Finished At (last testcase)",
    ]
    for col, h in enumerate(report_headers, start=1):
        c = ws_reports.cell(row=1, column=col, value=h)
        c.font = bold
        c.alignment = Alignment(horizontal="center")

    for idx, r in enumerate(reports, start=2):
        finished_dt = last_finished_map.get(r.id) or getattr(r, "finished", None)
        ws_reports.cell(row=idx, column=1, value=getattr(r, "report_id", None) or f"R{r.id}")
        ws_reports.cell(row=idx, column=2, value=getattr(r, "triggered_in", "") or "")
        ws_reports.cell(row=idx, column=3, value=str(getattr(r, "triggered_by", "") or "") or "")
        ws_reports.cell(row=idx, column=4, value=int(getattr(r, "total_passed", 0) or 0))
        ws_reports.cell(row=idx, column=5, value=int(getattr(r, "total_failed", 0) or 0))
        ws_reports.cell(row=idx, column=6, value=int(getattr(r, "total_blocked", 0) or 0))
        ws_reports.cell(row=idx, column=7, value=(getattr(r, "started", None).strftime("%Y-%m-%d %H:%M") if getattr(r, "started", None) else "—"))
        ws_reports.cell(row=idx, column=8, value=(finished_dt.strftime("%Y-%m-%d %H:%M") if finished_dt else "—"))

    ws_reports.freeze_panes = ws_reports["A2"]
    ws_reports.column_dimensions["A"].width = 12
    ws_reports.column_dimensions["B"].width = 40
    ws_reports.column_dimensions["C"].width = 22
    ws_reports.column_dimensions["D"].width = 8
    ws_reports.column_dimensions["E"].width = 8
    ws_reports.column_dimensions["F"].width = 8
    ws_reports.column_dimensions["G"].width = 18
    ws_reports.column_dimensions["H"].width = 22

    # --- Testcases sheet ---
    tc_headers = [
        "Automation Report ID",
        "Testcase ID",
        "Request Name",
        "Run ID",
        "Outcome",
        "Status",
        "Started At",
        "Finished At",
    ]
    for col, h in enumerate(tc_headers, start=1):
        c = ws_testcases.cell(row=1, column=col, value=h)
        c.font = bold
        c.alignment = Alignment(horizontal="center")

    for idx, t in enumerate(testcase_rows, start=2):
        ws_testcases.cell(row=idx, column=1, value=t.get("automation_report_id"))
        ws_testcases.cell(row=idx, column=2, value=t.get("testcase_id") or "")
        ws_testcases.cell(row=idx, column=3, value=t.get("request_name") or "")
        ws_testcases.cell(row=idx, column=4, value=t.get("run_id") or "")
        ws_testcases.cell(row=idx, column=5, value=t.get("outcome") or "")
        ws_testcases.cell(row=idx, column=6, value=t.get("status") or "")
        sa = t.get("started_at")
        fa = t.get("finished_at")
        ws_testcases.cell(row=idx, column=7, value=(sa.strftime("%Y-%m-%d %H:%M") if sa else "—"))
        ws_testcases.cell(row=idx, column=8, value=(fa.strftime("%Y-%m-%d %H:%M") if fa else "—"))

    ws_testcases.freeze_panes = ws_testcases["A2"]
    ws_testcases.column_dimensions["A"].width = 20
    ws_testcases.column_dimensions["B"].width = 14
    ws_testcases.column_dimensions["C"].width = 30
    ws_testcases.column_dimensions["D"].width = 10
    ws_testcases.column_dimensions["E"].width = 12
    ws_testcases.column_dimensions["F"].width = 10
    ws_testcases.column_dimensions["G"].width = 18
    ws_testcases.column_dimensions["H"].width = 18

    # Serialize workbook to response
    from io import BytesIO

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    filename = f"automated_reports_{start_str or 'all'}_to_{end_str or 'all'}.xlsx".replace(":", "-")
    resp = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


@login_required
def automation_testcase_reports_export(request):
    """Export Test Case Reports to an Excel file.

    Query params:
      - start: YYYY-MM-DD (optional)
      - end: YYYY-MM-DD (optional)

    Summary groups counts by day based on finished timestamp (updated_at).
    """

    if account_models:
        _log_user_action(request, account_models.UserAuditTrail.Actions.EXPORT_TESTCASE_REPORT)

    start_str = (request.GET.get("start") or "").strip()
    end_str = (request.GET.get("end") or "").strip()
    start_date = None
    end_date = None
    try:
        if start_str:
            start_date = timezone.datetime.fromisoformat(start_str).date()
    except Exception:
        start_date = None
    try:
        if end_str:
            end_date = timezone.datetime.fromisoformat(end_str).date()
    except Exception:
        end_date = None

    qs = models.ApiRunResultReport.objects.select_related(
        "automation_report",
        "run",
        "request",
        "testcase",
        "automation_report__triggered_by",
    ).order_by("-updated_at", "-id")

    # Filter by finished date (updated_at) when provided
    if start_date:
        qs = qs.filter(updated_at__date__gte=start_date)
    if end_date:
        qs = qs.filter(updated_at__date__lte=end_date)

    rows = list(qs)

    # Summary buckets by day
    summary_by_day: dict[str, dict[str, int]] = {}
    for obj in rows:
        try:
            day_key = obj.updated_at.date().isoformat() if getattr(obj, "updated_at", None) else "Unfinished"
        except Exception:
            day_key = "Unfinished"

        bucket = summary_by_day.setdefault(day_key, {"passed": 0, "failed": 0, "blocked": 0, "total": 0})
        try:
            status_norm = str(getattr(obj, "status", "") or "").strip().lower()
        except Exception:
            status_norm = ""

        if status_norm == "passed":
            bucket["passed"] += 1
        elif status_norm == "failed":
            bucket["failed"] += 1
        elif status_norm == "error":
            bucket["blocked"] += 1
        else:
            # Unexpected statuses count as blocked so totals still reconcile.
            bucket["blocked"] += 1
        bucket["total"] += 1

    try:
        from openpyxl import Workbook
        from openpyxl.chart import LineChart, Reference
        from openpyxl.styles import Alignment, Font
    except Exception as exc:
        return HttpResponse(f"Excel export dependency not available: {exc}", status=500)

    wb = Workbook()
    try:
        wb.remove(wb.active)
    except Exception:
        pass

    ws_summary = wb.create_sheet("Summary")
    ws_details = wb.create_sheet("Testcases")

    bold = Font(bold=True)

    exported_at = timezone.now()
    range_label = f"{start_str or '—'} to {end_str or '—'}"
    ws_summary["A1"].value = "Test Case Report Export"
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary["A2"].value = "Date range"
    ws_summary["A2"].font = bold
    ws_summary["B2"].value = range_label
    ws_summary["A3"].value = "Exported at"
    ws_summary["A3"].font = bold
    ws_summary["B3"].value = exported_at.strftime("%Y-%m-%d %H:%M")

    header_row = 5
    headers = ["Day", "Passed", "Failed", "Blocked", "Total"]
    for col, h in enumerate(headers, start=1):
        c = ws_summary.cell(row=header_row, column=col, value=h)
        c.font = bold
        c.alignment = Alignment(horizontal="center")

    day_keys = sorted([k for k in summary_by_day.keys() if k != "Unfinished"])
    if "Unfinished" in summary_by_day:
        day_keys.append("Unfinished")

    r0 = header_row + 1
    for i, day in enumerate(day_keys):
        b = summary_by_day[day]
        ws_summary.cell(row=r0 + i, column=1, value=day)
        ws_summary.cell(row=r0 + i, column=2, value=b["passed"])
        ws_summary.cell(row=r0 + i, column=3, value=b["failed"])
        ws_summary.cell(row=r0 + i, column=4, value=b["blocked"])
        ws_summary.cell(row=r0 + i, column=5, value=b["total"])

    ws_summary.freeze_panes = ws_summary["A6"]

    if day_keys:
        chart = LineChart()
        chart.title = "Testcases by Day"
        chart.y_axis.title = "Count"
        chart.x_axis.title = "Day"
        data = Reference(ws_summary, min_col=2, max_col=4, min_row=header_row, max_row=r0 + len(day_keys) - 1)
        cats = Reference(ws_summary, min_col=1, min_row=r0, max_row=r0 + len(day_keys) - 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        for s in chart.series:
            try:
                s.marker.symbol = "circle"
                s.marker.size = 7
            except Exception:
                pass
            try:
                s.graphicalProperties.line.width = 20000
            except Exception:
                pass
        ws_summary.add_chart(chart, "G5")

    ws_summary.column_dimensions["A"].width = 14
    ws_summary.column_dimensions["B"].width = 10
    ws_summary.column_dimensions["C"].width = 10
    ws_summary.column_dimensions["D"].width = 10
    ws_summary.column_dimensions["E"].width = 10
    ws_summary.column_dimensions["G"].width = 18

    # Details sheet
    d_headers = [
        "Automation Report ID",
        "Report ID",
        "Triggered In",
        "Triggered By",
        "Testcase ID",
        "Request Name",
        "Run ID",
        "Status",
        "Started At",
        "Finished At",
    ]
    for col, h in enumerate(d_headers, start=1):
        c = ws_details.cell(row=1, column=col, value=h)
        c.font = bold
        c.alignment = Alignment(horizontal="center")

    for idx, obj in enumerate(rows, start=2):
        ar = getattr(obj, "automation_report", None)
        ws_details.cell(row=idx, column=1, value=getattr(obj, "automation_report_id", None))
        ws_details.cell(row=idx, column=2, value=(getattr(ar, "report_id", "") if ar else ""))
        ws_details.cell(row=idx, column=3, value=(getattr(ar, "triggered_in", "") if ar else ""))
        ws_details.cell(row=idx, column=4, value=str(getattr(getattr(ar, "triggered_by", None), "username", "") or getattr(ar, "triggered_by", "") or "") if ar else "")
        ws_details.cell(row=idx, column=5, value=(obj.testcase.testcase_id if getattr(obj, "testcase", None) else ""))
        ws_details.cell(row=idx, column=6, value=(obj.request.name if getattr(obj, "request", None) else ""))
        ws_details.cell(row=idx, column=7, value=(getattr(obj, "run_id", None) or ""))
        ws_details.cell(row=idx, column=8, value=(getattr(obj, "status", "") or ""))
        sa = getattr(obj, "created_at", None)
        fa = getattr(obj, "updated_at", None)
        ws_details.cell(row=idx, column=9, value=(sa.strftime("%Y-%m-%d %H:%M") if sa else "—"))
        ws_details.cell(row=idx, column=10, value=(fa.strftime("%Y-%m-%d %H:%M") if fa else "—"))

    ws_details.freeze_panes = ws_details["A2"]
    ws_details.column_dimensions["A"].width = 20
    ws_details.column_dimensions["B"].width = 12
    ws_details.column_dimensions["C"].width = 40
    ws_details.column_dimensions["D"].width = 22
    ws_details.column_dimensions["E"].width = 14
    ws_details.column_dimensions["F"].width = 30
    ws_details.column_dimensions["G"].width = 10
    ws_details.column_dimensions["H"].width = 10
    ws_details.column_dimensions["I"].width = 18
    ws_details.column_dimensions["J"].width = 18

    from io import BytesIO

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    filename = f"testcase_reports_{start_str or 'all'}_to_{end_str or 'all'}.xlsx".replace(":", "-")
    resp = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


@ensure_csrf_cookie
@login_required
def automation_projects(request):
    data = _prepare_automation_data()
    context = {
        "initial_plans": data["plans"],
        "initial_metrics": data["metrics"],
        "api_endpoints": data["api_endpoints"],
    }
    return render(request, "automation/projects/automation_projects.html", context)


@ensure_csrf_cookie
@login_required
def automation_project_scenarios(request):
    data = _prepare_automation_data()
    # include initial modules so the scenarios page can show module filters/selects
    context = {
        "initial_plans": data["plans"],
        "initial_metrics": data["metrics"],
        "api_endpoints": data["api_endpoints"],
        "initial_selected_plan": data["selected_plan"],
        "initial_selected_scenario": data["selected_scenario"],
    }
    try:
        context["initial_modules"] = serializers.TestModulesSerializer(
            models.TestModules.objects.all(), many=True
        ).data
    except Exception:
        context["initial_modules"] = []
    return render(request, "automation/scenarios/automation_project_scenarios.html", context)


@ensure_csrf_cookie
@login_required
def automation_test_cases(request):
    data = _prepare_automation_data()
    selected_plan = data["selected_plan"]
    selected_scenario = data["selected_scenario"]

    plan_param = request.GET.get("plan")
    scenario_param = request.GET.get("scenario")

    scenario_match = None
    plan_match = None

    if scenario_param not in (None, ""):
        try:
            scenario_id = int(scenario_param)
        except (TypeError, ValueError):
            scenario_id = None
        if scenario_id is not None:
            for plan in data.get("plans", []):
                for scenario in plan.get("scenarios", []):
                    try:
                        if int(scenario.get("id")) == scenario_id:
                            scenario_match = scenario
                            plan_match = plan
                            break
                    except (TypeError, ValueError):
                        continue
                if scenario_match:
                    break
    if scenario_match:
        selected_scenario = scenario_match
        selected_plan = plan_match or selected_plan
    elif plan_param not in (None, ""):
        try:
            project_id = int(plan_param)
        except (TypeError, ValueError):
            project_id = None
        if project_id is not None:
            for plan in data.get("plans", []):
                try:
                    if int(plan.get("id")) == project_id:
                        selected_plan = plan
                        scenarios = plan.get("scenarios", []) or []
                        selected_scenario = scenarios[0] if scenarios else None
                        break
                except (TypeError, ValueError):
                    continue
    if scenario_param in (None, "") and plan_param in (None, ""):
        selected_plan = None
        selected_scenario = None

    context = {
        "initial_plans": data["plans"],
        "initial_metrics": data["metrics"],
        "api_endpoints": data["api_endpoints"],
        "initial_selected_plan": selected_plan,
        "initial_selected_scenario": selected_scenario,
    }
    # include initial modules so the test cases page can populate modal selects
    try:
        context["initial_modules"] = serializers.TestModulesSerializer(
            models.TestModules.objects.all(), many=True
        ).data
    except Exception:
        context["initial_modules"] = []
    return render(request, "automation/testcases/automation_test_cases.html", context)


@ensure_csrf_cookie
@login_required
def automation_test_plan_maintenance(request):
    data = _prepare_automation_data()
    context = {
        "initial_plans": data["plans"],
        "initial_metrics": data["metrics"],
        "api_endpoints": data["api_endpoints"],
        "initial_selected_plan": data["selected_plan"],
        "initial_selected_scenario": data["selected_scenario"],
    }
    return render(request, "core/automation_test_plan_maintenance.html", context)


@ensure_csrf_cookie
@login_required
def automation_data_management(request, section: str | None = None):
    data = _prepare_automation_data()
    context = {
        "initial_metrics": data["metrics"],
        "initial_environments": data["environments"],
        "initial_risks": data["risks"],
        "initial_mitigation_plans": data["mitigation_plans"],
        "initial_risk_mitigations": data["risk_mitigations"],
        "api_endpoints": data["api_endpoints"],
        "initial_section": section or "",
    }
    return render(request, "core/automation_data_management.html", context)


@ensure_csrf_cookie
@login_required
def automation_data_management_api_environment(request, section: str | None = None):
    """Render API Environments focused data management page."""
    data = _prepare_automation_data()
    context = {
        "initial_metrics": data["metrics"],
        "initial_environments": data["environments"],
        "initial_risks": data["risks"],
        "initial_mitigation_plans": data["mitigation_plans"],
        "initial_risk_mitigations": data["risk_mitigations"],
        "api_endpoints": data["api_endpoints"],
        "initial_section": section or "environments",
    }
    return render(request, "core/automation_data_management_api_environment.html", context)


@ensure_csrf_cookie
@login_required
def automation_data_management_mitigation_plan(request, section: str | None = None):
    """Render mitigation plan focused data management page."""
    data = _prepare_automation_data()
    context = {
        "initial_metrics": data["metrics"],
        "initial_environments": data["environments"],
        "initial_risks": data["risks"],
        "initial_mitigation_plans": data["mitigation_plans"],
        "initial_risk_mitigations": data["risk_mitigations"],
        "api_endpoints": data["api_endpoints"],
        "initial_section": section or "mitigation",
    }
    return render(request, "core/automation_data_management_mitigation_plan.html", context)


@ensure_csrf_cookie
@login_required
def automation_data_management_risk_registry(request, section: str | None = None):
    data = _prepare_automation_data()
    context = {
        "initial_metrics": data["metrics"],
        "initial_environments": data["environments"],
        "initial_risks": data["risks"],
        "initial_mitigation_plans": data["mitigation_plans"],
        "initial_risk_mitigations": data["risk_mitigations"],
        "api_endpoints": data["api_endpoints"],
        "initial_section": section or "",
    }
    return render(request, "core/automation_data_management_risk_registry.html", context)


@ensure_csrf_cookie
@login_required
def automation_data_management_test_tools(request, section: str | None = None):
    """Render Test Tools focused data management page."""
    data = _prepare_automation_data()
    context = {
        "initial_metrics": data["metrics"],
        "initial_environments": data["environments"],
        "initial_risks": data["risks"],
        "initial_mitigation_plans": data["mitigation_plans"],
        "initial_risk_mitigations": data["risk_mitigations"],
        "api_endpoints": data["api_endpoints"],
        "initial_section": section or "test-tools",
    }
    return render(request, "core/automation_data_management_test_tools.html", context)


@ensure_csrf_cookie
@login_required
def automation_project_modules(request, section: str | None = None):
    """Render Test Modules focused data management page."""
    data = _prepare_automation_data()
    context = {
        "initial_metrics": data["metrics"],
        "initial_environments": data["environments"],
        "initial_risks": data["risks"],
        "initial_mitigation_plans": data["mitigation_plans"],
        "initial_risk_mitigations": data["risk_mitigations"],
        "initial_plans": data.get("plans", []),
        "initial_test_modules": data.get("test_modules", []),
        "api_endpoints": data["api_endpoints"],
        "initial_section": section or "test-modules",
    }
    return render(request, "automation/modules/automation_project_modules.html", context)



class ApiAdhocRequestView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):  # noqa: D401
        payload = request.data or {}

        if account_models:
            action = account_models.UserAuditTrail.Actions.RUN_TEST_CASE
            report_id = payload.get("automation_report_id") or payload.get("automationReportId")
            if report_id:
                try:
                    report = models.AutomationReport.objects.filter(id=int(report_id)).only("triggered_in").first()
                    if report and report.triggered_in:
                        triggered_text = str(report.triggered_in).lower()
                        if any(token in triggered_text for token in ("project", "module", "scenario", "case", "testcase", "individual", "run")):
                            action = account_models.UserAuditTrail.Actions.RUN_AUTOMATION_TEST_CASE
                except Exception:
                    pass

            if action != account_models.UserAuditTrail.Actions.RUN_AUTOMATION_TEST_CASE:
                _log_user_action(request, action)

        # Temporary debug: log a safe summary of incoming body_transforms and environment
        try:
            transforms = payload.get("body_transforms")
            if isinstance(transforms, dict):
                summary = []
                for ov in transforms.get("overrides", []) or []:
                    try:
                        summary.append(
                            {
                                "path": ov.get("path"),
                                "has_value": bool(ov.get("value") not in (None, "")),
                                "isRandom": bool(ov.get("isRandom") or ov.get("is_random")),
                                "charLimitProvided": ov.get("charLimit") is not None or ov.get("char_limit") is not None,
                            }
                        )
                    except Exception:
                        # ignore per-override errors
                        continue
                logger.info("execute payload.transforms summary: %s", json.dumps(summary))
        except Exception:
            logger.exception("failed to summarize incoming body_transforms")

        try:
            logger.info("execute payload.environment: %s", payload.get("environment"))
        except Exception:
            pass

        method = str(payload.get("method", "GET")).upper()
        url = payload.get("url")
        if not url:
            raise ValidationError({"url": "URL is required."})

        headers = payload.get("headers") or {}
        if not isinstance(headers, dict):
            raise ValidationError({"headers": "Headers must be an object."})

        params = payload.get("params") or {}
        if not isinstance(params, dict):
            raise ValidationError({"params": "Query params must be an object."})

        overrides = payload.get("overrides") or {}
        if not isinstance(overrides, dict):
            raise ValidationError({"overrides": "Overrides must be an object."})

        timeout = payload.get("timeout", 30)
        try:
            timeout = float(timeout)
        except (TypeError, ValueError):
            raise ValidationError({"timeout": "Timeout must be numeric."})

        environment = None
        variables: dict[str, Any] = {}
        environment_id = payload.get("environment")
        # If an explicit environment is provided, accept either an integer id or an environment name.
        if environment_id is not None and environment_id != "":
            # Try numeric pk first
            try:
                environment = models.ApiEnvironment.objects.filter(pk=int(environment_id)).first()
            except Exception:
                environment = None

            # If numeric lookup failed, try to match by environment name (case-insensitive)
            if environment is None:
                try:
                    environment = models.ApiEnvironment.objects.filter(name__iexact=str(environment_id)).first()
                except Exception:
                    environment = None

            if environment:
                variables.update(environment.variables or {})
                default_headers = environment.default_headers or {}
                headers = {**default_headers, **headers}

        collection = None
        collection_id = payload.get("collection_id")
        if collection_id not in (None, ""):
            try:
                collection_id = int(collection_id)
            except (TypeError, ValueError) as exc:
                raise ValidationError({"collection": "Collection must be a valid integer."}) from exc
            try:
                collection = models.ApiCollection.objects.prefetch_related('environments').get(pk=collection_id)
            except models.ApiCollection.DoesNotExist as exc:
                raise ValidationError({"collection": "Collection not found."}) from exc

        api_request = None
        request_id = payload.get("request_id")
        if request_id not in (None, ""):
            try:
                request_id = int(request_id)
            except (TypeError, ValueError) as exc:
                raise ValidationError({"request": "Request must be a valid integer."}) from exc
            api_request = models.ApiRequest.objects.select_related("collection").filter(pk=request_id).first()
            if api_request:
                if collection and api_request.collection_id != collection.id:
                    api_request = None
                elif not collection:
                    collection = api_request.collection

        # If no explicit environment was provided but the collection has environments,
        # prefer an environment that contains any template variables referenced by
        # the request's body_transforms (for example `non_realtime_mid`). If none
        # match, fall back to the first environment.
        if environment is None and collection is not None:
            try:
                transforms_to_apply = None
                if isinstance(payload.get("body_transforms"), dict):
                    transforms_to_apply = payload.get("body_transforms")
                elif api_request and getattr(api_request, "body_transforms", None):
                    transforms_to_apply = api_request.body_transforms

                chosen_env = None
                envs_qs = collection.environments.all()
                # If transforms reference template variables, try to pick an env that contains them
                if transforms_to_apply:
                    try:
                        import json as _json
                        raw_text = _json.dumps(transforms_to_apply)
                        keys = set(services.VARIABLE_PATTERN.findall(raw_text))
                        if keys:
                            # Prefer an environment that contains all referenced keys (strong match)
                            for candidate in envs_qs:
                                vars_map = candidate.variables or {}
                                if all(k in vars_map for k in keys):
                                    chosen_env = candidate
                                    break
                            # If no env contains all keys, fall back to any env that has at least one of the keys
                            if chosen_env is None:
                                for candidate in envs_qs:
                                    vars_map = candidate.variables or {}
                                    if any(k in vars_map for k in keys):
                                        chosen_env = candidate
                                        break
                    except Exception:
                        chosen_env = None

                if chosen_env is None:
                    chosen_env = envs_qs.first()

                if chosen_env is not None:
                    environment = chosen_env
                    variables.update(environment.variables or {})
                    default_headers = environment.default_headers or {}
                    headers = {**default_headers, **headers}
            except Exception:
                # ignore failures to read collection environments and continue
                pass

        # Finally, merge any runtime overrides which should take precedence over environment vars
        variables.update(overrides)

        if overrides:
            try:
                overrides_snapshot = json.dumps(overrides, ensure_ascii=False)
            except TypeError:
                overrides_snapshot = str(overrides)
            truncated_snapshot = overrides_snapshot[:2000]
            logger.info("[tester.execute] runtime overrides: %s", truncated_snapshot)
            print("[tester.execute] runtime overrides:", truncated_snapshot)
            try:
                for key, value in overrides.items():
                    normalized = str(key).lower()
                    if "pay" in normalized or "reference" in normalized:
                        logger.info("[tester.execute] override %s=%s", key, value)
                        print(f"[tester.execute] override {key}={value}")
            except Exception:
                pass

        # Resolve url/headers/params after collection/environment and overrides have been merged
        resolved_url = services._resolve_variables(url, variables)  # type: ignore[attr-defined]
        resolved_headers = services._resolve_variables(headers, variables)  # type: ignore[attr-defined]
        resolved_params = services._resolve_variables(params, variables)  # type: ignore[attr-defined]

        form_data_entries = payload.get("form_data") or []
        if form_data_entries and not isinstance(form_data_entries, list):
            raise ValidationError({"form_data": "form_data must be an array."})

        json_body = payload.get("json")
        body = payload.get("body")
        resolved_json = None
        resolved_body: Any = None
        files_payload: dict[str, tuple[str, io.BytesIO, str]] | None = None
        request_form_snapshot: list[dict[str, Any]] | None = None

        if form_data_entries:
            text_fields: dict[str, Any] = {}
            file_fields: dict[str, dict[str, Any]] = {}
            request_form_snapshot = []

            for entry in form_data_entries:
                if not isinstance(entry, dict):
                    continue
                key = str(entry.get("key", "")).strip()
                if not key:
                    continue
                entry_type = entry.get("type", "text")
                if entry_type == "file":
                    data_url = entry.get("data")
                    if not data_url or not isinstance(data_url, str):
                        continue
                    try:
                        header, encoded = data_url.split(",", 1)
                    except ValueError as exc:  # pragma: no cover - malformed payload
                        raise ValidationError({"form_data": f"Invalid file data for '{key}'."}) from exc
                    if ";base64" not in header:
                        raise ValidationError({"form_data": f"File data for '{key}' must be base64 encoded."})
                    try:
                        file_bytes = base64.b64decode(encoded)
                    except (binascii.Error, ValueError) as exc:  # pragma: no cover - malformed payload
                        raise ValidationError({"form_data": f"File data for '{key}' is not valid base64."}) from exc
                    filename = entry.get("filename") or "upload.bin"
                    content_type = entry.get("content_type") or "application/octet-stream"
                    file_fields[key] = {
                        "filename": filename,
                        "bytes": file_bytes,
                        "content_type": content_type,
                    }
                    request_form_snapshot.append(
                        {
                            "key": key,
                            "type": "file",
                            "filename": filename,
                            "content_type": content_type,
                            "size": len(file_bytes),
                        }
                    )
                else:
                    value = entry.get("value", "")
                    if not isinstance(value, str):
                        value = str(value)
                    text_fields[key] = value
                    request_form_snapshot.append({"key": key, "type": "text", "value": value})

            if text_fields:
                resolved_text = services._resolve_variables(text_fields, variables)  # type: ignore[attr-defined]
                if isinstance(resolved_text, dict):
                    resolved_body = resolved_text
                else:  # pragma: no cover - defensive fallback
                    resolved_body = text_fields
                for snapshot_entry in request_form_snapshot:
                    if snapshot_entry.get("type") == "text":
                        snapshot_entry["value"] = resolved_body.get(snapshot_entry["key"], snapshot_entry.get("value", ""))
            elif request_form_snapshot:  # at least one file entry
                resolved_body = None

            if file_fields:
                files_payload = {
                    key: (meta["filename"], io.BytesIO(meta["bytes"]), meta["content_type"])
                    for key, meta in file_fields.items()
                }
        elif json_body is not None:
            if not isinstance(json_body, (dict, list)):
                raise ValidationError({"json": "JSON body must be an object or array."})
            resolved_json = services._resolve_variables(json_body, variables)  # type: ignore[attr-defined]
            # If JSON payload is an object, apply body_transforms (overrides/signatures).
            if isinstance(resolved_json, dict):
                transforms_to_apply = None
                if isinstance(payload.get("body_transforms"), dict):
                    transforms_to_apply = payload.get("body_transforms")
                elif api_request and getattr(api_request, "body_transforms", None):
                    transforms_to_apply = api_request.body_transforms
                if transforms_to_apply:
                    try:
                        overrides_map = services._apply_body_transforms(resolved_json, transforms_to_apply, variables)
                        # update variables with any values produced by signature builders
                        variables.update(overrides_map or {})
                    except Exception:
                        # best-effort: continue with untransformed payload
                        pass
        elif body not in (None, ""):
            if isinstance(body, (dict, list)):
                resolved_json = services._resolve_variables(body, variables)  # type: ignore[attr-defined]
            else:
                resolved_body = services._resolve_variables(str(body), variables)  # type: ignore[attr-defined]
                # If body is raw XML and transforms are present, apply XML transforms
                try:
                    raw_type = payload.get("body_raw_type") or ""
                    if isinstance(raw_type, str) and raw_type.lower() == "xml":
                        transforms_to_apply = None
                        if isinstance(payload.get("body_transforms"), dict):
                            transforms_to_apply = payload.get("body_transforms")
                        elif api_request and getattr(api_request, "body_transforms", None):
                            transforms_to_apply = api_request.body_transforms
                        if transforms_to_apply:
                            try:
                                resolved_body = services._apply_xml_body_transforms(resolved_body, transforms_to_apply, variables)
                            except Exception:
                                pass
                except Exception:
                    pass

        signature_value = None
        if isinstance(resolved_json, dict):
            signature_value = resolved_json.get("signature")
        elif isinstance(resolved_body, dict):
            signature_value = resolved_body.get("signature")
        elif isinstance(resolved_body, str):
            try:
                parsed_body = json.loads(resolved_body)
                if isinstance(parsed_body, dict):
                    signature_value = parsed_body.get("signature")
            except (TypeError, ValueError):
                pass

        if signature_value not in (None, ""):
            logger.info("API tester resolved signature: %s", signature_value)

        run = models.ApiRun.objects.create(
            collection=collection,
            environment=environment,
            triggered_by=request.user if request.user.is_authenticated else None,
            status=models.ApiRun.Status.RUNNING,
            started_at=timezone.now(),
        )
        run_result = models.ApiRunResult.objects.create(
            run=run,
            request=api_request,
            order=1,
            status=models.ApiRunResult.Status.ERROR,
        )
        # track any AutomationReport created during execution so we can
        # include its id in the response payload to clients
        automation_report = None

        outbound_plaintext = None
        outbound_payload = None
        payload_reencrypted = False
        if isinstance(resolved_json, dict):
            try:
                outbound_plaintext, outbound_payload, payload_reencrypted = _apply_pay_reference_override(resolved_json, overrides)
                if payload_reencrypted:
                    logger.info("[tester.execute] outbound payload updated with overridden pay_reference.")
                    print("[tester.execute] outbound payload updated with overridden pay_reference.")
                    if outbound_payload is not None:
                        try:
                            outbound_preview = json.dumps(outbound_payload, ensure_ascii=False)[:2000]
                        except TypeError:
                            outbound_preview = str(outbound_payload)[:2000]
                        logger.info("[tester.execute] outbound decrypted payload: %s", outbound_preview)
                        print("[tester.execute] outbound decrypted payload:", outbound_preview)
            except Exception:
                # helper already logs; continue without blocking execution
                pass

        start = time.perf_counter()
        try:
            if resolved_json is not None:
                try:
                    logger.info(
                        "API tester outbound JSON: %s",
                        json.dumps(resolved_json, ensure_ascii=False)[:2000],
                    )
                except Exception:
                    logger.info("API tester outbound JSON (unserializable)")
            elif resolved_body not in (None, ""):
                body_preview = resolved_body
                if isinstance(body_preview, str) and len(body_preview) > 2000:
                    body_preview = f"{body_preview[:2000]}…"
                logger.info("API tester outbound body: %s", body_preview)
            response = requests.request(
                method=method,
                url=resolved_url,
                headers=resolved_headers,
                params=resolved_params,
                data=None if resolved_json is not None else resolved_body,
                json=resolved_json,
                files=files_payload,
                timeout=max(1.0, float(timeout)),
            )
        except requests.RequestException as exc:  # pragma: no cover - network error path
            elapsed_ms = (time.perf_counter() - start) * 1000
            run_result.error = str(exc)
            run_result.response_time_ms = elapsed_ms
            run_result.status = models.ApiRunResult.Status.ERROR
            run_result.save(update_fields=["error", "response_time_ms", "status", "updated_at"])
            run.status = models.ApiRun.Status.FAILED
            run.summary = services._summarize_run(1, 0)  # type: ignore[attr-defined]
            run.finished_at = timezone.now()
            run.save(update_fields=["status", "summary", "finished_at", "updated_at"])
            # mirror into report table (non-blocking)
            try:
                tc = None
                try:
                    tc = api_request.test_cases.first()
                except Exception:
                    tc = None

                automation_report = None
                # Prefer an explicit automation_report_id from the client payload
                try:
                    ar_id = payload.get("automation_report_id") if isinstance(payload, dict) else None
                except Exception:
                    ar_id = None
                if ar_id:
                    try:
                        automation_report = models.AutomationReport.objects.filter(pk=int(ar_id)).first()
                    except Exception:
                        automation_report = None
                # Fallback to finding/creating by run.started_at and triggered_by
                if not automation_report:
                    try:
                        automation_report = models.AutomationReport.objects.filter(started=run.started_at, triggered_by=run.triggered_by).first()
                        if not automation_report:
                            automation_report = models.AutomationReport.objects.create(
                                triggered_in=(run.collection.name if run.collection else ""),
                                triggered_by=run.triggered_by,
                                started=run.started_at,
                                finished=run.finished_at,
                            )
                    except Exception:
                        automation_report = None

                models.ApiRunResultReport.objects.create(
                    run=run,
                    request=api_request,
                    order=run_result.order,
                    status=run_result.status,
                    response_status=run_result.response_status,
                    response_headers=run_result.response_headers,
                    response_body=run_result.response_body,
                    response_time_ms=run_result.response_time_ms,
                    assertions_passed=run_result.assertions_passed,
                    assertions_failed=run_result.assertions_failed,
                    error=run_result.error,
                    testcase=tc,
                    automation_report=automation_report,
                )
                # recompute report totals based on test case results
                try:
                    services.recompute_automation_report_totals(automation_report)
                    if automation_report is not None:
                        # ensure finished is current
                        if run.finished_at and (not automation_report.finished or run.finished_at > automation_report.finished):
                            automation_report.finished = run.finished_at
                            automation_report.save(update_fields=["finished"])
                except Exception:
                    pass
            except Exception:
                pass
            payload = {
                "error": str(exc),
                "resolved_url": resolved_url,
                "request_headers": resolved_headers,
                "run_id": run.id,
            }
            try:
                if automation_report is not None:
                    payload["automation_report_id"] = getattr(automation_report, "id", None)
            except Exception:
                pass
            return Response(payload, status=status.HTTP_502_BAD_GATEWAY)

        elapsed_ms = (time.perf_counter() - start) * 1000

        try:
            response_json = response.json()
        except ValueError:
            response_json = None

        if response_json is not None:
            try:
                response_snapshot = json.dumps(response_json, ensure_ascii=False)
            except TypeError:
                response_snapshot = str(response_json)
            truncated_response = response_snapshot[:2000]
            logger.info("[tester.execute] response json: %s", truncated_response)
            print("[tester.execute] response json:", truncated_response)

            pay_reference = None
            decrypted_plaintext = None
            decrypted_payload: Any | None = None

            if isinstance(response_json, dict):
                try:
                    pay_reference = response_json.get("pay_reference")
                    if not pay_reference and isinstance(response_json.get("data"), dict):
                        pay_reference = response_json["data"].get("pay_reference")
                except Exception:
                    pay_reference = None

                encrypted_field = response_json.get("data")
                if isinstance(encrypted_field, str):
                    decrypted_plaintext, decrypted_payload = _attempt_decrypt_response_data(encrypted_field)
                    if decrypted_plaintext:
                        truncated_plaintext = decrypted_plaintext[:2000]
                        logger.info("[tester.execute] decrypted text: %s", truncated_plaintext)
                        print("[tester.execute] decrypted text:", truncated_plaintext)
                    if decrypted_payload is not None:
                        try:
                            decrypted_snapshot = json.dumps(decrypted_payload, ensure_ascii=False)
                        except TypeError:
                            decrypted_snapshot = str(decrypted_payload)
                        truncated_decrypted = decrypted_snapshot[:2000]
                        logger.info("[tester.execute] decrypted json: %s", truncated_decrypted)
                        print("[tester.execute] decrypted json:", truncated_decrypted)
                        if pay_reference is None and isinstance(decrypted_payload, dict):
                            pay_reference = decrypted_payload.get("pay_reference")

            if pay_reference:
                logger.info("[tester.execute] response pay_reference=%s", pay_reference)
                print(f"[tester.execute] response pay_reference={pay_reference}")

        run_result.response_status = response.status_code
        run_result.response_headers = dict(response.headers)
        run_result.response_body = response.text[:20000]
        run_result.response_time_ms = elapsed_ms
        run_result.status = models.ApiRunResult.Status.PASSED if response.ok else models.ApiRunResult.Status.FAILED
        run_result.save(
            update_fields=[
                "response_status",
                "response_headers",
                "response_body",
                "response_time_ms",
                "status",
                "updated_at",
            ]
        )
        # mirror into report table (non-blocking)
        try:
            tc = None
            try:
                tc = api_request.test_cases.first()
            except Exception:
                tc = None

            automation_report = None
            try:
                ar_id = payload.get("automation_report_id") if isinstance(payload, dict) else None
            except Exception:
                ar_id = None
            if ar_id:
                try:
                    automation_report = models.AutomationReport.objects.filter(pk=int(ar_id)).first()
                except Exception:
                    automation_report = None
            if not automation_report:
                try:
                    # try to find an existing report for this run
                    automation_report = models.AutomationReport.objects.filter(started=run.started_at, triggered_by=run.triggered_by).first()
                    if not automation_report:
                        # fallback: create a new report with collection name as triggered_in
                        automation_report = models.AutomationReport.objects.create(
                            triggered_in=(run.collection.name if run.collection else ""),
                            triggered_by=run.triggered_by,
                            started=run.started_at,
                        )
                except Exception:
                    automation_report = None

            models.ApiRunResultReport.objects.create(
                run=run,
                request=api_request,
                order=run_result.order,
                status=run_result.status,
                response_status=run_result.response_status,
                response_headers=run_result.response_headers,
                response_body=run_result.response_body,
                response_time_ms=run_result.response_time_ms,
                assertions_passed=run_result.assertions_passed,
                assertions_failed=run_result.assertions_failed,
                error=run_result.error,
                testcase=tc,
                automation_report=automation_report,
            )
            # recompute totals based on TestCase latest results and update finished
            try:
                services.recompute_automation_report_totals(automation_report)
                if automation_report is not None:
                    if run.finished_at and (not automation_report.finished or run.finished_at > automation_report.finished):
                        automation_report.finished = run.finished_at
                        automation_report.save(update_fields=["finished"])
            except Exception:
                pass
        except Exception:
            pass
        passed = 1 if response.ok else 0
        run.status = models.ApiRun.Status.PASSED if response.ok else models.ApiRun.Status.FAILED
        run.summary = services._summarize_run(1, passed)  # type: ignore[attr-defined]
        run.finished_at = timezone.now()
        run.save(update_fields=["status", "summary", "finished_at", "updated_at"])

        return Response(
            {
                "status_code": response.status_code,
                "headers": dict(response.headers),
                "body": response.text[:20000],
                "json": response_json,
                "elapsed_ms": elapsed_ms,
                "resolved_url": resolved_url,
                "request": {
                    "method": method,
                    "headers": resolved_headers,
                    "params": resolved_params,
                    "body": resolved_body,
                    "json": resolved_json,
                    "form_data": request_form_snapshot or None,
                    "timeout": timeout,
                },
                "environment": environment.name if environment else None,
                "variables": variables,
                "run_id": run.id,
                "run_result_id": run_result.id,
                "automation_report_id": (automation_report.id if automation_report is not None else None),
            }
        )


@login_required
def api_tester_page(request):
    """Render the interactive API testing workspace."""
    return render(request, "core/api_tester.html")


class AutomationReportFinalizeView(APIView):
    """Endpoint to recompute totals and mark an AutomationReport finished.

    Expects JSON body: {"report_id": <int>, "finished": "ISO timestamp (optional)"}
    """
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        report_id = request.data.get("report_id")
        if not report_id:
            return Response({"report_id": "report_id is required"}, status=status.HTTP_400_BAD_REQUEST)
        try:
            report = models.AutomationReport.objects.get(pk=int(report_id))
        except models.AutomationReport.DoesNotExist:
            raise NotFound("AutomationReport not found")

        # Allow client to provide final totals (useful for blocked/skipped cases
        # that do not produce server-side report rows). If `totals` is present
        # we persist these values directly. Otherwise recompute from stored
        # ApiRunResultReport rows.
        totals = request.data.get("totals")
        if isinstance(totals, dict):
            try:
                passed = int(totals.get("passed") or 0)
            except Exception:
                passed = 0
            try:
                failed = int(totals.get("failed") or 0)
            except Exception:
                failed = 0
            try:
                blocked = int(totals.get("blocked") or 0)
            except Exception:
                blocked = 0
            try:
                report.total_passed = max(0, passed)
                report.total_failed = max(0, failed)
                report.total_blocked = max(0, blocked)
                report.save(update_fields=["total_passed", "total_failed", "total_blocked"])
            except Exception as exc:  # pragma: no cover - defensive
                logger.exception("Failed to persist provided totals for AutomationReport %s: %s", report_id, exc)
                return Response({"error": "failed to persist provided totals"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        else:
            try:
                services.recompute_automation_report_totals(report)
            except Exception as exc:
                logger.exception("Failed to recompute totals for AutomationReport %s: %s", report_id, exc)
                return Response({"error": "failed to recompute totals"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        finished_val = request.data.get("finished")
        if finished_val:
            try:
                # try ISO parse, fallback to now
                finished_dt = timezone.datetime.fromisoformat(str(finished_val))
                if timezone.is_naive(finished_dt):
                    finished_dt = timezone.make_aware(finished_dt, timezone.get_current_timezone())
            except Exception:
                finished_dt = timezone.now()
        else:
            finished_dt = timezone.now()

        try:
            if not report.finished or finished_dt > report.finished:
                report.finished = finished_dt
                report.save(update_fields=["finished"])
        except Exception as exc:  # pragma: no cover - defensive
            logger.exception("Failed to save finished timestamp for AutomationReport %s: %s", report_id, exc)

        try:
            serializer = serializers.AutomationReportSerializer(report)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception:
            return Response({"status": "ok"}, status=status.HTTP_200_OK)


class AutomationReportCreateView(APIView):
    """Create a new AutomationReport for a batch run.

    Expects optional JSON: {"triggered_in": "<string>", "started": "ISO (optional)"}
    Returns serialized AutomationReport.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        try:
            logger.info("[automation] AutomationReport create called by user=%s payload=%s", getattr(request, 'user', None), request.data)
        except Exception:
            pass
        triggered_in = request.data.get("triggered_in") or request.data.get("triggeredIn") or ""
        started_val = request.data.get("started")
        try:
            if started_val:
                try:
                    started_dt = timezone.datetime.fromisoformat(str(started_val))
                    if timezone.is_naive(started_dt):
                        started_dt = timezone.make_aware(started_dt, timezone.get_current_timezone())
                except Exception:
                    started_dt = timezone.now()
            else:
                started_dt = timezone.now()
            report = models.AutomationReport.objects.create(
                triggered_in=str(triggered_in)[:500],
                triggered_by=request.user if request.user and request.user.is_authenticated else None,
                started=started_dt,
            )
            if account_models:
                action = _automation_run_action(triggered_in)
                if action:
                    _log_user_action(request, action)
            serializer = serializers.AutomationReportSerializer(report)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        except Exception as exc:
            logger.exception("Failed to create AutomationReport: %s", exc)
            return Response({"error": "failed to create report"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class AutomationReportDetailView(APIView):
    """Detail endpoint for AutomationReport allowing partial updates (PATCH).

    URL: /api/core/automation-report/<pk>/
    Accepts PATCH with any of: `total_passed`|`total_success`|`passed`,
    `total_failed`|`failed`, `total_blocked`|`blocked`, and optional `finished`.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, pk=None, *args, **kwargs):
        """Return serialized AutomationReport for the given pk.

        This allows clients to refresh canonical totals after a finalize.
        """
        try:
            report = models.AutomationReport.objects.get(pk=int(pk))
        except (ValueError, models.AutomationReport.DoesNotExist):
            raise NotFound("AutomationReport not found")

        serializer = serializers.AutomationReportSerializer(report, context={})
        return Response(serializer.data, status=status.HTTP_200_OK)

    def patch(self, request, pk=None, *args, **kwargs):
        try:
            report = models.AutomationReport.objects.get(pk=int(pk))
        except (ValueError, models.AutomationReport.DoesNotExist):
            raise NotFound("AutomationReport not found")

        data = request.data or {}
        try:
            logger.info("[automation] PATCH automation-report payload: %s", data)
        except Exception:
            pass

        # Normalise totals from several possible client keys
        def parse_int(keys):
            for k in keys:
                if k in data:
                    try:
                        return int(data.get(k) or 0)
                    except Exception:
                        return 0
            return None

        updated_fields = []
        passed = parse_int(["total_passed", "total_success", "passed"])
        failed = parse_int(["total_failed", "failed"])
        blocked = parse_int(["total_blocked", "blocked"])
        try:
            logger.info(
                "[automation] PATCH parsed totals -> passed=%s failed=%s blocked=%s",
                passed,
                failed,
                blocked,
            )
        except Exception:
            pass

        if passed is not None:
            report.total_passed = max(0, passed)
            updated_fields.append("total_passed")
        if failed is not None:
            report.total_failed = max(0, failed)
            updated_fields.append("total_failed")
        if blocked is not None:
            report.total_blocked = max(0, blocked)
            updated_fields.append("total_blocked")
        try:
            report.save(update_fields=updated_fields or None)
        except Exception:
            try:
                report.save()
            except Exception:
                pass
        serializer = serializers.AutomationReportSerializer(report)
        return Response(serializer.data, status=status.HTTP_200_OK)


class AutomationReportTestcaseDetailView(APIView):
    """Return a single ApiRunResultReport for an AutomationReport and testcase id.

    URL: /api/core/automation-report/<report_pk>/testcase/<testcase_id>/
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, pk=None, testcase_id=None, *args, **kwargs):
        try:
            # Try to find by linked TestCase.testcase_id
            qs = models.ApiRunResultReport.objects.select_related("testcase", "run", "request")
            obj = qs.filter(automation_report_id=int(pk), testcase__testcase_id=str(testcase_id)).order_by("-created_at").first()
            if not obj:
                # fallback: try matching by TestCase PK if numeric
                try:
                    if str(testcase_id).isdigit():
                        obj = qs.filter(automation_report_id=int(pk), testcase_id=int(testcase_id)).order_by("-created_at").first()
                except Exception:
                    obj = None
            if not obj:
                raise models.ApiRunResultReport.DoesNotExist()
        except (ValueError, models.ApiRunResultReport.DoesNotExist):
            raise NotFound("Testcase report not found")

        serializer = serializers.ApiRunResultReportSerializer(obj, context={})
        return Response(serializer.data, status=status.HTTP_200_OK)

        # Allow client to set finished timestamp (ISO string) or use now
        finished_val = data.get("finished")
        if finished_val is not None:
            try:
                finished_dt = timezone.datetime.fromisoformat(str(finished_val))
                if timezone.is_naive(finished_dt):
                    finished_dt = timezone.make_aware(finished_dt, timezone.get_current_timezone())
            except Exception:
                finished_dt = timezone.now()
            report.finished = finished_dt
            updated_fields.append("finished")

        try:
            if updated_fields:
                report.save(update_fields=updated_fields)
        except Exception as exc:
            logger.exception("Failed to patch AutomationReport %s: %s", pk, exc)
            return Response({"error": "failed to update report"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        serializer = serializers.AutomationReportSerializer(report, context={})
        return Response(serializer.data, status=status.HTTP_200_OK)


def _pid_is_running(pid: int | None) -> bool:
    if not pid or pid <= 0:
        return False
    try:
        os.kill(int(pid), 0)
        return True
    except OSError:
        return False


def _safe_mkdir(path: Path) -> None:
    try:
        path.mkdir(parents=True, exist_ok=True)
    except Exception:
        # Let callers fail later when writing files.
        pass


def _loadtest_collect_testcases(*, scope: str, selection: dict[str, Any]) -> list[models.TestCase]:
    """Return testcases from the hierarchy selection.

    Selection payload keys (all optional): project_ids, module_ids, scenario_ids, testcase_ids.
    """

    qs = models.TestCase.objects.select_related("scenario", "related_api_request")
    qs = qs.filter(scenario__is_automated=True)

    def as_int_list(value: Any) -> list[int]:
        if value in (None, ""):
            return []
        if isinstance(value, (list, tuple)):
            raw = list(value)
        else:
            raw = [value]
        out: list[int] = []
        for v in raw:
            try:
                out.append(int(v))
            except Exception:
                continue
        return out

    if scope == models.LoadTestRun.Scope.TESTCASE:
        ids = as_int_list(selection.get("testcase_ids"))
        if ids:
            qs = qs.filter(pk__in=ids)
    elif scope == models.LoadTestRun.Scope.SCENARIO:
        ids = as_int_list(selection.get("scenario_ids"))
        if ids:
            qs = qs.filter(scenario_id__in=ids)
    elif scope == models.LoadTestRun.Scope.MODULE:
        ids = as_int_list(selection.get("module_ids"))
        if ids:
            qs = qs.filter(scenario__module_id__in=ids)
    elif scope == models.LoadTestRun.Scope.PROJECT:
        ids = as_int_list(selection.get("project_ids"))
        if ids:
            qs = qs.filter(scenario__project_id__in=ids)

    # Only include testcases that have a related request.
    qs = qs.filter(related_api_request__isnull=False)
    return list(qs.order_by("scenario_id", "id"))


def _loadtest_build_execute_payloads(*, testcases: list[models.TestCase], environment_id: int | None) -> list[dict[str, Any]]:
    payloads: list[dict[str, Any]] = []
    for tc in testcases:
        req = getattr(tc, "related_api_request", None)
        if not req:
            continue
        payload: dict[str, Any] = {
            "request_id": req.id,
            "collection_id": req.collection_id,
            "method": req.method or "GET",
            "url": req.url or "",
            "headers": req.headers or {},
            "params": req.query_params or {},
            "timeout": max(1, float(getattr(req, "timeout_ms", 30000) or 30000) / 1000.0),
        }
        if environment_id is not None:
            payload["environment"] = int(environment_id)

        # Attach body content per request config
        body_type = str(getattr(req, "body_type", "") or "").lower()
        if body_type == "json":
            payload["json"] = req.body_json or {}
        elif body_type == "form":
            form_entries = []
            try:
                if isinstance(req.body_form, dict):
                    for k, v in (req.body_form or {}).items():
                        form_entries.append({"key": k, "type": "text", "value": v})
            except Exception:
                form_entries = []
            payload["form_data"] = form_entries
        elif body_type == "raw":
            if req.body_raw:
                payload["body"] = req.body_raw
            if getattr(req, "body_raw_type", None):
                payload["body_raw_type"] = req.body_raw_type

        # Attach transforms so the execute endpoint can apply them
        if getattr(req, "body_transforms", None):
            payload["body_transforms"] = req.body_transforms

        # Authorization headers for configured request auth
        try:
            auth_type = str(getattr(req, "auth_type", "") or "").lower()
            if auth_type == "bearer" and getattr(req, "auth_bearer", ""):
                payload["headers"] = payload.get("headers") or {}
                payload["headers"]["Authorization"] = f"Bearer {req.auth_bearer}"
            elif auth_type == "basic" and isinstance(getattr(req, "auth_basic", None), dict):
                ab = req.auth_basic or {}
                username = ab.get("username") or ""
                password = ab.get("password") or ""
                token = base64.b64encode(f"{username}:{password}".encode("utf-8")).decode("utf-8")
                payload["headers"] = payload.get("headers") or {}
                payload["headers"]["Authorization"] = f"Basic {token}"
        except Exception:
            pass

        # Helpful label for Locust stats grouping
        payload["_locust_name"] = f"{tc.testcase_id or tc.id} · {req.name or req.id}"
        payloads.append(payload)
    return payloads


def _write_locustfile(*, dest: Path, auth_token: str, payloads: list[dict[str, Any]], host: str) -> None:
    """Write a locustfile that posts to our own tester execute endpoint."""

    safe_payloads = []
    for p in payloads:
        # Ensure non-serializable values won't break locustfile creation.
        try:
            json.dumps(p)
            safe_payloads.append(p)
        except TypeError:
            safe_payloads.append({k: v for k, v in p.items() if isinstance(v, (str, int, float, bool, type(None), dict, list))})

    # IMPORTANT: embed payloads as JSON text and parse with json.loads.
    # This avoids invalid Python literals like `null`, `true`, `false`.
    payloads_json_text = json.dumps(safe_payloads, ensure_ascii=False)

    content = """# Auto-generated by Automation Load Testing\n\nimport json\nimport random\nfrom locust import HttpUser, task, constant\n\nAUTH_TOKEN = {auth_token!r}\nHOST = {host!r}\nPAYLOADS = json.loads({payloads_json_text!r})\n\n\nclass AutomationLoadTestUser(HttpUser):\n    host = HOST\n    wait_time = constant(0)\n\n    @task\n    def execute_testcase(self):\n        payload = random.choice(PAYLOADS)\n        name = payload.get('_locust_name') or 'tester.execute'\n        headers = {{\n            'Authorization': f'Token {{AUTH_TOKEN}}',\n            'Content-Type': 'application/json',\n            'Accept': 'application/json',\n        }}\n        # The server returns 200 even when the upstream request fails; treat\n        # execute HTTP errors as failures, and also treat execute responses that\n        # carry an `error` field as failures for the Locust report.\n        with self.client.post('/api/core/tester/execute/', json=payload, headers=headers, name=name, catch_response=True) as resp:\n            if not resp.ok:\n                resp.failure(f'execute failed HTTP {{resp.status_code}}')\n                return\n            try:\n                data = resp.json()\n            except Exception:\n                data = None\n            if isinstance(data, dict) and data.get('error'):\n                resp.failure(str(data.get('error'))[:300])\n                return\n            resp.success()\n""".format(
        auth_token=auth_token,
        host=host,
        payloads_json_text=payloads_json_text,
    )
    dest.write_text(content, encoding="utf-8")


def _tail_text_file(path: Path, *, max_bytes: int = 8000) -> str:
    try:
        if not path.exists() or not path.is_file():
            return ""
        size = path.stat().st_size
        start = max(0, size - max_bytes)
        with open(path, "rb") as f:
            f.seek(start)
            data = f.read(max_bytes)
        text = data.decode("utf-8", errors="replace")
        return text.strip()
    except Exception:
        return ""


def _revoke_knox_token(token_key: str) -> None:
    if not token_key or not AuthToken:
        return
    try:
        AuthToken.objects.filter(token_key=token_key).delete()
    except Exception:
        pass


class LoadTestRunsApiView(APIView):
    """Create and list LoadTestRun records."""

    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        media_url = getattr(settings, "MEDIA_URL", None) or "/media/"
        qs = models.LoadTestRun.objects.select_related("created_by").order_by("-created_at")[:50]

        # Refresh status for stale RUNNING rows (e.g., locust crashed early).
        try:
            media_root = Path(settings.MEDIA_ROOT or "media")
        except Exception:
            media_root = Path("media")

        for obj in qs:
            if obj.status != models.LoadTestRun.Status.RUNNING:
                continue
            if _pid_is_running(obj.locust_pid):
                continue
            # Locust exited; mark as finished or error based on exit_code/log.
            log_tail = ""
            try:
                if obj.log_relpath:
                    log_tail = _tail_text_file(media_root / obj.log_relpath)
            except Exception:
                log_tail = ""
            # If locust crashed, the log will contain a traceback.
            new_status = models.LoadTestRun.Status.FINISHED
            if obj.exit_code not in (None, 0):
                new_status = models.LoadTestRun.Status.ERROR
            if "Traceback" in log_tail or "Error" in log_tail:
                new_status = models.LoadTestRun.Status.ERROR
            obj.status = new_status
            obj.finished_at = obj.finished_at or timezone.now()
            if new_status == models.LoadTestRun.Status.ERROR and not obj.error:
                obj.error = (log_tail.splitlines()[-1] if log_tail else "Locust exited with an error")[:2000]
            try:
                obj.save(update_fields=["status", "finished_at", "error", "updated_at"])
            except Exception:
                pass
            _revoke_knox_token(obj.knox_token_key)

        items = []
        for obj in qs:
            items.append(
                {
                    "id": obj.id,
                    "name": obj.name,
                    "status": obj.status,
                    "scope": obj.scope,
                    "selection": obj.selection,
                    "users": obj.users,
                    "ramp_up_seconds": obj.ramp_up_seconds,
                    "duration_seconds": obj.duration_seconds,
                    "spawn_rate": obj.spawn_rate,
                    "started_at": obj.started_at.isoformat() if obj.started_at else None,
                    "finished_at": obj.finished_at.isoformat() if obj.finished_at else None,
                    "pid": obj.locust_pid,
                    "exit_code": obj.exit_code,
                    "report_html": (media_url + obj.report_html_relpath) if obj.report_html_relpath else None,
                    "csv_prefix": (media_url + obj.csv_prefix_relpath) if obj.csv_prefix_relpath else None,
                    "log": (media_url + obj.log_relpath) if obj.log_relpath else None,
                    "error": obj.error,
                }
            )
        return Response(items, status=status.HTTP_200_OK)

    def post(self, request, *args, **kwargs):
        if not AuthToken:
            return Response({"error": "Knox is not available"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        data = request.data or {}
        scope = str(data.get("scope") or models.LoadTestRun.Scope.TESTCASE).strip().lower()
        if scope not in {
            models.LoadTestRun.Scope.PROJECT,
            models.LoadTestRun.Scope.MODULE,
            models.LoadTestRun.Scope.SCENARIO,
            models.LoadTestRun.Scope.TESTCASE,
        }:
            return Response({"scope": "Invalid scope"}, status=status.HTTP_400_BAD_REQUEST)

        selection = data.get("selection") if isinstance(data.get("selection"), dict) else {}

        # Config
        try:
            users = int(data.get("users") or 1)
        except Exception:
            users = 1
        users = max(1, min(users, 50000))

        try:
            ramp_up_seconds = int(data.get("ramp_up_seconds") or 0)
        except Exception:
            ramp_up_seconds = 0
        ramp_up_seconds = max(0, ramp_up_seconds)

        # Optional: explicit spawn rate (users/sec). If provided, it takes
        # precedence over ramp_up_seconds.
        spawn_rate_override = data.get("spawn_rate")
        spawn_rate: float | None
        try:
            spawn_rate = float(spawn_rate_override) if spawn_rate_override not in (None, "") else None
        except Exception:
            spawn_rate = None
        if spawn_rate is not None and spawn_rate <= 0:
            spawn_rate = None

        try:
            duration_seconds = int(data.get("duration_seconds") or 60)
        except Exception:
            duration_seconds = 60
        duration_seconds = max(1, duration_seconds)

        if spawn_rate is None:
            # Ensure ramp-up window isn't longer than the test itself; otherwise
            # Locust may never reach the requested user count.
            if ramp_up_seconds > duration_seconds:
                ramp_up_seconds = duration_seconds
            # Derived spawn rate (users per second)
            spawn_rate = float(users) if ramp_up_seconds <= 0 else max(0.1, float(users) / float(ramp_up_seconds))
        else:
            # Derive ramp seconds for recordkeeping.
            try:
                ramp_up_seconds = int(max(0, min(duration_seconds, int((float(users) / float(spawn_rate)) + 0.9999))))
            except Exception:
                ramp_up_seconds = 0

        environment_id = None

        name = str(data.get("name") or "").strip()[:180]

        run = models.LoadTestRun.objects.create(
            name=name,
            status=models.LoadTestRun.Status.CREATED,
            created_by=(request.user if request.user and request.user.is_authenticated else None),
            scope=scope,
            selection=selection,
            users=users,
            ramp_up_seconds=ramp_up_seconds,
            duration_seconds=duration_seconds,
            spawn_rate=spawn_rate,
        )

        # Resolve testcases
        testcases = _loadtest_collect_testcases(scope=scope, selection=selection)
        if not testcases:
            run.status = models.LoadTestRun.Status.FAILED
            run.error = "No automated test cases found for selection (or no related API request configured)."
            run.finished_at = timezone.now()
            run.save(update_fields=["status", "error", "finished_at", "updated_at"])
            return Response({"error": run.error, "id": run.id}, status=status.HTTP_400_BAD_REQUEST)

        payloads = _loadtest_build_execute_payloads(testcases=testcases, environment_id=environment_id)
        if not payloads:
            run.status = models.LoadTestRun.Status.FAILED
            run.error = "No executable payloads could be built (missing related API requests)."
            run.finished_at = timezone.now()
            run.save(update_fields=["status", "error", "finished_at", "updated_at"])
            return Response({"error": run.error, "id": run.id}, status=status.HTTP_400_BAD_REQUEST)

        # Work directory under MEDIA_ROOT
        media_root = Path(settings.MEDIA_ROOT or "media")
        workdir = media_root / "load_tests" / f"run_{run.id}"
        _safe_mkdir(workdir)

        # Locust artifacts
        locustfile = workdir / "locustfile.py"
        csv_prefix = workdir / "stats"
        report_html = workdir / "report.html"
        log_file = workdir / "locust.log"

        # Mint a token for this run (revoked after completion/stop)
        token_obj, token = AuthToken.objects.create(request.user)
        run.knox_token_key = getattr(token_obj, "token_key", "") or ""

        # Determine our own host for locust to call into.
        # Use the current request host (works both local and in docker).
        # If the app is behind a proxy, ensure Django is configured for correct scheme.
        override_host = os.environ.get("AUTOMATION_LOADTEST_HOST")
        scheme = "https" if request.is_secure() else "http"
        host = str(override_host).strip() if override_host else f"{scheme}://{request.get_host()}"
        try:
            parsed = urlparse(host)
            if not parsed.scheme or not parsed.netloc:
                host = "http://localhost:8000"
        except Exception:
            host = "http://localhost:8000"

        try:
            _write_locustfile(dest=locustfile, auth_token=token, payloads=payloads, host=host)
        except Exception as exc:
            _revoke_knox_token(run.knox_token_key)
            run.status = models.LoadTestRun.Status.FAILED
            run.error = f"Failed to write locustfile: {exc}"
            run.finished_at = timezone.now()
            run.save(update_fields=["status", "error", "finished_at", "updated_at", "knox_token_key"])
            return Response({"error": run.error, "id": run.id}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # Build locust command (headless)
        cmd = [
            sys.executable,
            "-m",
            "locust",
            "-f",
            str(locustfile),
            "--headless",
            "--users",
            str(users),
            "--spawn-rate",
            str(spawn_rate),
            "--run-time",
            f"{duration_seconds}s",
            "--csv",
            str(csv_prefix),
            "--html",
            str(report_html),
            "--only-summary",
        ]

        try:
            with open(log_file, "w", encoding="utf-8") as lf:
                proc = subprocess.Popen(
                    cmd,
                    cwd=str(workdir),
                    stdout=lf,
                    stderr=subprocess.STDOUT,
                    env={**os.environ},
                )
        except Exception as exc:
            _revoke_knox_token(run.knox_token_key)
            run.status = models.LoadTestRun.Status.FAILED
            run.error = f"Failed to start locust: {exc}"
            run.finished_at = timezone.now()
            run.save(update_fields=["status", "error", "finished_at", "updated_at", "knox_token_key"])
            return Response({"error": run.error, "id": run.id}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        run.status = models.LoadTestRun.Status.RUNNING
        run.started_at = timezone.now()
        run.locust_pid = int(proc.pid)
        run.workdir = str(workdir)
        run.report_html_relpath = str(report_html.relative_to(media_root))
        # Prefix (csv) is multiple files; store directory/prefix relative for UI links
        run.csv_prefix_relpath = str(csv_prefix.relative_to(media_root))
        run.log_relpath = str(log_file.relative_to(media_root))
        run.save(
            update_fields=[
                "status",
                "started_at",
                "locust_pid",
                "workdir",
                "report_html_relpath",
                "csv_prefix_relpath",
                "log_relpath",
                "knox_token_key",
                "updated_at",
            ]
        )

        def _waiter(run_id: int, proc_handle: subprocess.Popen, token_key: str, log_path: Path):
            exit_code: int | None = None
            try:
                # Wait for locust to exit. Add buffer to allow html/csv flush.
                exit_code = proc_handle.wait(timeout=float(duration_seconds) + 120.0)
            except Exception:
                # If it didn't exit, just fall back to pid polling.
                try:
                    for _ in range(0, duration_seconds + 3600):
                        if not _pid_is_running(proc_handle.pid):
                            break
                        time.sleep(1)
                except Exception:
                    pass
                try:
                    exit_code = proc_handle.poll()
                except Exception:
                    exit_code = None

            try:
                obj = models.LoadTestRun.objects.filter(pk=run_id).first()
                if not obj:
                    return
                if obj.status == models.LoadTestRun.Status.STOPPED:
                    return
                # Determine final status
                tail = _tail_text_file(log_path)
                final_status = models.LoadTestRun.Status.FINISHED
                if exit_code not in (None, 0):
                    final_status = models.LoadTestRun.Status.ERROR
                if "Traceback" in tail or "NameError" in tail or "Exception" in tail:
                    final_status = models.LoadTestRun.Status.ERROR

                obj.exit_code = exit_code
                obj.status = final_status
                obj.finished_at = timezone.now()
                if final_status == models.LoadTestRun.Status.ERROR:
                    # keep a short summary
                    if tail:
                        last_line = tail.splitlines()[-1]
                        obj.error = (last_line or "Locust exited with an error")[:2000]
                    else:
                        obj.error = "Locust exited with an error"
                obj.save(update_fields=["status", "exit_code", "finished_at", "error", "updated_at"])
            finally:
                _revoke_knox_token(token_key)

        try:
            t = threading.Thread(
                target=_waiter,
                args=(run.id, proc, run.knox_token_key, log_file),
                daemon=True,
            )
            t.start()
        except Exception:
            pass

        return Response({"id": run.id, "status": run.status}, status=status.HTTP_201_CREATED)


class LoadTestRunDetailApiView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk=None, *args, **kwargs):
        media_url = getattr(settings, "MEDIA_URL", None) or "/media/"
        try:
            obj = models.LoadTestRun.objects.get(pk=int(pk))
        except (ValueError, models.LoadTestRun.DoesNotExist):
            raise NotFound("Load test run not found")

        # Refresh running state if PID is dead.
        if obj.status == models.LoadTestRun.Status.RUNNING and not _pid_is_running(obj.locust_pid):
            obj.status = models.LoadTestRun.Status.FINISHED
            obj.finished_at = obj.finished_at or timezone.now()
            obj.save(update_fields=["status", "finished_at", "updated_at"])
            _revoke_knox_token(obj.knox_token_key)

        payload = {
            "id": obj.id,
            "name": obj.name,
            "status": obj.status,
            "scope": obj.scope,
            "selection": obj.selection,
            "users": obj.users,
            "ramp_up_seconds": obj.ramp_up_seconds,
            "duration_seconds": obj.duration_seconds,
            "spawn_rate": obj.spawn_rate,
            "started_at": obj.started_at.isoformat() if obj.started_at else None,
            "finished_at": obj.finished_at.isoformat() if obj.finished_at else None,
            "pid": obj.locust_pid,
            "exit_code": obj.exit_code,
            "report_html": (media_url + obj.report_html_relpath) if obj.report_html_relpath else None,
            "csv_prefix": (media_url + obj.csv_prefix_relpath) if obj.csv_prefix_relpath else None,
            "log": (media_url + obj.log_relpath) if obj.log_relpath else None,
            "error": obj.error,
        }
        return Response(payload, status=status.HTTP_200_OK)


class LoadTestRunStopApiView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, pk=None, *args, **kwargs):
        try:
            obj = models.LoadTestRun.objects.get(pk=int(pk))
        except (ValueError, models.LoadTestRun.DoesNotExist):
            raise NotFound("Load test run not found")

        pid = obj.locust_pid
        if obj.status != models.LoadTestRun.Status.RUNNING or not _pid_is_running(pid):
            # ensure consistent state
            if obj.status == models.LoadTestRun.Status.RUNNING:
                obj.status = models.LoadTestRun.Status.FINISHED
                obj.finished_at = obj.finished_at or timezone.now()
                obj.save(update_fields=["status", "finished_at", "updated_at"])
            _revoke_knox_token(obj.knox_token_key)
            return Response({"status": obj.status}, status=status.HTTP_200_OK)

        try:
            os.kill(int(pid), signal.SIGTERM)
        except Exception as exc:
            return Response({"error": f"Failed to stop locust: {exc}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        obj.status = models.LoadTestRun.Status.STOPPED
        obj.finished_at = timezone.now()
        obj.save(update_fields=["status", "finished_at", "updated_at"])
        _revoke_knox_token(obj.knox_token_key)
        return Response({"status": obj.status}, status=status.HTTP_200_OK)
